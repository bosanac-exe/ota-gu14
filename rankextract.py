import concurrent.futures
import os
import re
import time
import pandas as pd
import requests
import urllib3
import openpyxl
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl.styles import Font, PatternFill, Alignment

# Suppress insecure request warnings caused by verify=False
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Target base and navigation URLs
base_url = "https://ota.tournamentsoftware.com"
landing_url = "https://ota.tournamentsoftware.com/ranking/ranking.aspx?rid=143"

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

session = requests.Session()
session.headers.update(headers)

# Set of club names that should be replaced with "No Club"
NO_CLUB_VARIATIONS = {
    "AB No Club",
    "ON OR No Club",
    "British Columbia No Club",
    "BC NO No Club",
    "NO CLUB",
    "OTA- No Club",
    "ON East No Club",
    "Lower Mainland - No Club",
    "AA-Membre dans aucun club / Not a member of a Club",
    "AA-Mon club n’apparaît pas / My club is not listed",
    "SK No Club",
    "PE No Club",
    "Not a member of a Club / Membre dans aucun club"
}


def parse_wl_record(record_str):
    if not record_str or "/" not in record_str:
        return None, None
    try:
        clean_str = re.sub(r"\(.*\)", "", record_str).strip()
        parts = clean_str.split("/")
        return int(parts[0].strip()), int(parts[1].strip())
    except (ValueError, IndexError):
        return None, None


def calculate_percentage(wins, losses):
    if wins is None or losses is None:
        return ""
    total = wins + losses
    return round(wins / total, 2) if total > 0 else 0.0


def scrape_profile_details(player_info):
    url = player_info["_profile_url"]

    player_info["Singles WTN"] = "unavailable"
    player_info["Career Matches"] = ""
    player_info["Singles W/L-Career"] = "unavailable"
    player_info["Singles W/L-YTD"] = "unavailable"
    player_info["Singles W/L Career %"] = ""
    player_info["Singles W/L-YTD %"] = ""

    if not url:
        return player_info

    max_retries = 3
    backoff_factor = 1.5

    for attempt in range(max_retries):
        try:
            time.sleep(0.1)
            res = session.get(url, timeout=10, verify=False)

            if res.status_code == 200:
                p_soup = BeautifulSoup(res.text, "html.parser")

                # 1. Singles WTN Extraction
                singles_title = p_soup.find(
                    "span", class_="tag-duo__title", string=re.compile(r"^Singles$", re.I)
                )
                if singles_title:
                    val_span = singles_title.find_next_sibling("span", class_="tag-duo__value")
                    if val_span:
                        wtn_match = re.search(r"([\d\.]+)", val_span.text)
                        if wtn_match:
                            player_info["Singles WTN"] = float(wtn_match.group(1))

                # 2. Singles Win/Loss Statistics (Targeting #tabStatsSingles)
                singles_tab = p_soup.find("div", id="tabStatsSingles")
                if singles_tab:
                    career_item = singles_tab.find("dt", class_="list__label", string=re.compile(r"^Career$", re.I))
                    if career_item:
                        val_div = career_item.find_next("span", class_="list__value-start")
                        if val_div:
                            player_info["Singles W/L-Career"] = re.sub(r"\(.*\)", "", val_div.text).strip()

                    ytd_item = singles_tab.find("dt", class_="list__label", string=re.compile(r"^This year$", re.I))
                    if ytd_item:
                        val_div = ytd_item.find_next("span", class_="list__value-start")
                        if val_div:
                            player_info["Singles W/L-YTD"] = re.sub(r"\(.*\)", "", val_div.text).strip()

                if player_info["Singles WTN"] != "unavailable" or player_info["Singles W/L-Career"] != "unavailable":
                    c_wins, c_losses = parse_wl_record(player_info["Singles W/L-Career"])
                    
                    if c_wins is not None and c_losses is not None:
                        player_info["Career Matches"] = c_wins + c_losses
                    
                    player_info["Singles W/L Career %"] = calculate_percentage(c_wins, c_losses)

                    y_wins, y_losses = parse_wl_record(player_info["Singles W/L-YTD"])
                    player_info["Singles W/L-YTD %"] = calculate_percentage(y_wins, y_losses)
                    break

        except (requests.RequestException, Exception):
            pass
        
        time.sleep(backoff_factor * (attempt + 1))

    return player_info


def update_trends_sheet(wb, sheet_names):
    """Compares the two most recent weekly sheets and writes to the 'trends' tab."""
    if len(sheet_names) < 2:
        return
    
    df_new = pd.read_excel("master.xlsx", sheet_name=sheet_names[-1])
    df_old = pd.read_excel("master.xlsx", sheet_name=sheet_names[-2])
    
    if "trends" in wb.sheetnames:
        wb.remove(wb["trends"])
    
    ws = wb.create_sheet("trends")
    
    headers = [
        'Rank', 'Player', 'Year of Birth', 'Points', 'Tournaments',
        'Province', 'Club', 'Singles WTN', 'Career Matches',
        'Singles W/L-Career', 'Singles W/L-YTD %'
    ]
    ws.append(headers)
    
    # Style Header Row for Trends
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    font_red = Font(color="FF0000", bold=True)
    font_green = Font(color="008000", bold=True)
    font_black = Font(color="000000")

    df_old_lookup = df_old.set_index('Player')

    for _, row in df_new.iterrows():
        p = row['Player']
        row_26 = df_old_lookup.loc[p] if p in df_old_lookup.index else None
        if isinstance(row_26, pd.DataFrame): 
            row_26 = row_26.iloc[0]
            
        # Rank comparison
        r27 = row['Rank']
        r26 = row_26['Rank'] if row_26 is not None else None
        if r26 is None:
            rank_val, rank_color = f"{r27} (New)", font_green
        elif r27 < r26:
            rank_val, rank_color = f"{r27} (▲{int(r26 - r27)})", font_green
        elif r27 > r26:
            rank_val, rank_color = f"{r27} (▼{int(r27 - r26)})", font_red
        else:
            rank_val, rank_color = str(r27), font_black
            
        # Points comparison
        p27 = row['Points']
        p26 = row_26['Points'] if row_26 is not None else None
        if p26 is None:
            points_val, points_color = f"{p27:.2f}", font_black
        else:
            p_diff = p27 - p26
            if p_diff > 0:
                points_val, points_color = f"{p27:.2f} (▲{p_diff:.2f})", font_green
            elif p_diff < 0:
                points_val, points_color = f"{p27:.2f} (▼{abs(p_diff):.2f})", font_red
            else:
                points_val, points_color = f"{p27:.2f}", font_black

        # Tournaments comparison
        t27 = row['Tournaments']
        t26 = row_26['Tournaments'] if row_26 is not None else None
        if t26 is None:
            t_val, t_color = str(t27), font_black
        else:
            t_diff = t27 - t26
            if t_diff > 0:
                t_val, t_color = f"{t27} (+{int(t_diff)})", font_green
            elif t_diff < 0:
                t_val, t_color = f"{t27} ({int(t_diff)})", font_red
            else:
                t_val, t_color = str(t27), font_black

        # WTN comparison
        try:
            wtn27 = float(row['Singles WTN'])
        except:
            wtn27 = None
        try:
            wtn26 = float(row_26['Singles WTN']) if row_26 is not None else None
        except:
            wtn26 = None

        if wtn27 is None:
            wtn_val, wtn_color = str(row['Singles WTN']), font_black
        elif wtn26 is None:
            wtn_val, wtn_color = f"{wtn27:.1f}", font_black
        else:
            wtn_diff = wtn27 - wtn26
            if wtn_diff < 0:
                wtn_val, wtn_color = f"{wtn27:.1f} (▲{abs(wtn_diff):.1f})", font_green
            elif wtn_diff > 0:
                wtn_val, wtn_color = f"{wtn27:.1f} (▼{wtn_diff:.1f})", font_red
            else:
                wtn_val, wtn_color = f"{wtn27:.1f}", font_black

        # YTD Win Rate % comparison
        wlytd27 = row['Singles W/L-YTD %']
        wlytd26 = row_26['Singles W/L-YTD %'] if row_26 is not None else None
        if pd.isna(wlytd27) or wlytd27 == "":
            wlytd_val, wlytd_color = "0%", font_black
        elif wlytd26 is None or pd.isna(wlytd26) or wlytd26 == "":
            wlytd_val, wlytd_color = f"{int(float(wlytd27)*100)}%", font_black
        else:
            wl_diff = float(wlytd27) - float(wlytd26)
            if wl_diff > 0:
                wlytd_val, wlytd_color = f"{int(float(wlytd27)*100)}% (▲{int(wl_diff*100)}%)", font_green
            elif wl_diff < 0:
                wlytd_val, wlytd_color = f"{int(float(wlytd27)*100)}% (▼{int(abs(wl_diff)*100)}%)", font_red
            else:
                wlytd_val, wlytd_color = f"{int(float(wlytd27)*100)}%", font_black

        row_data = [
            rank_val, row['Player'], row['Year of Birth'], points_val, t_val,
            row['Province'], row['Club'], wtn_val, row['Career Matches'],
            row['Singles W/L-Career'], wlytd_val
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        ws.cell(row=curr_row, column=1).font = rank_color
        ws.cell(row=curr_row, column=4).font = points_color
        ws.cell(row=curr_row, column=5).font = t_color
        ws.cell(row=curr_row, column=8).font = wtn_color
        ws.cell(row=curr_row, column=11).font = wlytd_color

    # Formatting trends sheet
    ws.freeze_panes = "A2"
    max_col_letter = openpyxl.utils.get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def main():
    print("--- Phase 0: Automatically Discovering Latest Ranking URL ---")
    try:
        landing_res = session.get(landing_url, timeout=15, verify=False)
        landing_res.raise_for_status()
    except requests.RequestException as e:
        print(f"CRITICAL ERROR: Failed to access landing directory page: {e}")
        return

    landing_soup = BeautifulSoup(landing_res.text, "html.parser")
    target_link = landing_soup.find("a", string=lambda text: text and "Girls Under 14 Singles" in text)

    if not target_link or "href" not in target_link.attrs:
        print("CRITICAL ERROR: Could not locate 'Girls Under 14 Singles' link in page source.")
        return

    href = target_link["href"]
    if href.startswith("/"):
        base_category_url = base_url + href
    elif href.startswith("category.aspx"):
        base_category_url = base_url + "/ranking/" + href
    else:
        base_category_url = base_url + "/" + href

    page_url_template = base_category_url + "&C1698FOC=&p={page_num}&ps=100"
    print(f"Successfully configured dynamic path target:\n-> {page_url_template.format(page_num=1)}\n")

    players_master_list = []
    ranking_week = "unknown_week"

    print("--- Phase 1: Scraping Paginated Index Pages ---")
    for page in range(1, 10):
        print(f"Reading index list page {page}/10...")
        try:
            url = page_url_template.format(page_num=page)
            res = session.get(url, timeout=15, verify=False)
            res.raise_for_status()
        except requests.RequestException as e:
            print(f"Error reading listing page {page}: {e}")
            continue

        soup = BeautifulSoup(res.text, "html.parser")

        if ranking_week == "unknown_week":
            page_text = soup.get_text()
            week_match = re.search(r"Canada National Ranking[^(]*\(([^)]+)\)", page_text, re.I)
            if week_match:
                ranking_week = week_match.group(1).replace("/", "-").strip()

        rows = soup.find_all("tr")
        for row in rows:
            tds = row.find_all("td")
            if len(tds) >= 12:
                points_td = row.find("td", class_="rankingpoints")
                if not points_td:
                    continue
                try:
                    rank = int(tds[0].text.strip())
                    yob = int(tds[6].text.replace("\xa0", "").strip())
                    player_name = tds[7].find("a").text.strip() if tds[7].find("a") else tds[7].text.strip()

                    profile_anchor = tds[8].find("a")
                    full_profile_url = ""
                    if profile_anchor and "href" in profile_anchor.attrs:
                        link_href = profile_anchor["href"]
                        if link_href.startswith("/"):
                            full_profile_url = base_url + link_href
                        elif link_href.startswith("player.aspx") or link_href.startswith("player-profile/"):
                            full_profile_url = base_url + "/ranking/" + link_href if "player.aspx" in link_href else base_url + "/" + link_href
                        else:
                            full_profile_url = base_url + "/" + link_href

                    points = float(points_td.text.strip())
                    tournaments = int(tds[10].text.strip())
                    province = tds[11].find("a").text.strip() if tds[11].find("a") else tds[11].text.strip()
                    
                    club = tds[12].find("a").text.strip() if tds[12].find("a") else tds[12].text.strip()
                    if club in NO_CLUB_VARIATIONS or club.upper() == "NO CLUB":
                        club = "No Club"

                    excel_link = f'=HYPERLINK("{full_profile_url}", "Link")' if full_profile_url else ""

                    players_master_list.append({
                        "Rank": rank,
                        "Year of Birth": yob,
                        "Player": player_name,
                        "Profile": excel_link,
                        "Points": points,
                        "Tournaments": tournaments,
                        "Province": province,
                        "Club": club,
                        "_profile_url": full_profile_url
                    })
                except (IndexError, AttributeError, ValueError):
                    continue

    total_players = len(players_master_list)
    print(f"\nDiscovered {total_players} players. Initializing profile scans...")
    print("\n--- Phase 2: Parallel Deep Scraping of Player Profiles ---")

    final_processed_players = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(scrape_profile_details, player): player for player in players_master_list}
        
        for future in tqdm(concurrent.futures.as_completed(futures), total=total_players, desc="Processing Profiles"):
            updated_player_data = future.result()
            updated_player_data.pop("_profile_url", None)
            final_processed_players.append(updated_player_data)

    if final_processed_players:
        df = pd.DataFrame(final_processed_players)
        df = df.sort_values(by="Rank").reset_index(drop=True)

        column_order = [
            "Rank", "Year of Birth", "Player", "Profile", "Points", "Tournaments", 
            "Province", "Club", "Singles WTN", "Career Matches", "Singles W/L-Career", 
            "Singles W/L-YTD", "Singles W/L Career %", "Singles W/L-YTD %"
        ]
        df = df[column_order]

        output_filename = "master.xlsx"
        sheet_name = f"Week {ranking_week}"

        # Load existing workbook or create new if it doesn't exist
        if os.path.exists(output_filename):
            wb = openpyxl.load_workbook(output_filename)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet if creating fresh workbook
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        # Create the weekly sheet using pandas writer inside openpyxl context
        with pd.ExcelWriter(output_filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Re-load workbook to apply openpyxl formatting styles cleanly
        wb = openpyxl.load_workbook(output_filename)
        worksheet = wb[sheet_name]

        # Freeze the header row
        worksheet.freeze_panes = "A2"

        # Bold Header Font & Background Fill matching theme
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # AutoFilter dropdown elements across the headers
        max_col_letter = openpyxl.utils.get_column_letter(len(column_order))
        worksheet.auto_filter.ref = f"A1:{max_col_letter}{len(df) + 1}"

        # Number Formats
        for cell in worksheet["E"][1:]: # Points 
            cell.number_format = "0.000"

        for cell in worksheet["F"][1:]: # Tournaments 
            cell.number_format = "0"

        for cell in worksheet["J"][1:]: # Career Matches 
            if cell.value != "":
                cell.number_format = "0"

        for cell in worksheet["M"][1:]: # Career % 
            if cell.value != "":
                cell.number_format = "0%"

        for cell in worksheet["N"][1:]: # YTD % 
            if cell.value != "":
                cell.number_format = "0%"

        # Columns padding layout adjustments
        for col in worksheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    val_str = str(cell.value)
                    val_len = 8 if val_str.startswith("=HYPERLINK") else len(val_str)
                    if val_len > max_len:
                        max_len = val_len
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 11)

        # --- Trigger Trends Update ---
        weekly_sheets = [s for s in wb.sheetnames if s.startswith("Week")]
        if len(weekly_sheets) >= 2:
            update_trends_sheet(wb, sorted(weekly_sheets))

        wb.save(output_filename)
        print(f"\n🎉 Extraction Complete! Data written to sheet '{sheet_name}' inside '{output_filename}' with trends updated.")
    else:
        print("No players were detected.")

if __name__ == "__main__":
    main()
