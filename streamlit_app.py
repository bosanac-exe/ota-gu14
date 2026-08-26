import re
import json
from io import BytesIO
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import List, Tuple
import subprocess
import zoneinfo
import io

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from google import genai
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing, Line, PolyLine, Rect, String
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

WEEK_SHEET_RE = re.compile(r"^Week\s+(\d{1,2})-(\d{4})$", re.IGNORECASE)
MASTER_WORKBOOK = Path(__file__).with_name("master.xlsx")
PLAYER_PLACEHOLDER = "Select a player..."
PLAYER_VIEWS = {"Player history", "Player Charts", "YTD Player trends", "AI Analysis"}
LOWER_IS_BETTER = {"Rank", "Singles WTN"}
PERCENT_METRICS = {"Singles W/L Career %", "Singles W/L-YTD %"}
DISPLAY_COLUMNS = [
    "Week Label", "Rank", "Points", "Tournaments",
    "Singles WTN", "Career Matches", "Singles W/L-Career",
    "Singles W/L-YTD", "Singles W/L Career %", "Singles W/L-YTD %",
]
LATEST_TRENDS_COLUMNS = ["Rank", "Points", "Tournaments", "Singles WTN", "Career Matches", "Singles W/L Career %", "Singles W/L-YTD %"]
NUMERIC_TREND_COLUMNS = LATEST_TRENDS_COLUMNS.copy()
OPTIONAL_CHART_METRICS = ["Rank", "Points", "Singles WTN", "Tournaments", "Career Matches", "Singles W/L Career %", "Singles W/L-YTD %"]
YTD_TREND_METRICS = LATEST_TRENDS_COLUMNS.copy()
TOP_MOVER_METRIC_COLUMNS = ["Points", "Singles WTN", "Tournaments", "Career Matches", "Singles W/L-YTD %"]
WTN_OBJECT_ID_RE = re.compile(r"^[0-9a-fA-F]{24}$")

PROVINCE_MAPPING = {
    "Tennis QuÃ©bec": "QuÃ©bec",
    "Ontario Tennis Association": "Ontario",
    "British Columbia": "British Columbia",
    "Tennis Alberta": "Alberta",
    "Tennis Saskatchewan": "Saskatchewan",
    "Tennis New Brunswick": "New Brunswick",
    "Tennis Nova Scotia": "Nova Scotia",
    "Tennis Manitoba": "Manitoba",
    "Tennis Prince Edward Island": "Prince Edward Island",
    "Tennis Newfoundland & Labrador": "Newfoundland & Labrador"
}

def normalize_province_name(value):
    if pd.isna(value):
        return value
    val_str = str(value).strip()
    return PROVINCE_MAPPING.get(val_str, val_str)


def format_custom_datetime(dt: datetime) -> str:
    month_day_year = dt.strftime("%B %d, %Y")
    time_str = dt.strftime("%I:%M%p").lower().lstrip("0")
    return f"{month_day_year} - {time_str}"


@st.cache_data(show_spinner=False)
def get_data_updated_text(workbook_path: Path, modified_ns: int) -> str:
    try:
        git_epoch_str = subprocess.check_output(
            ["git", "log", "-1", "--format=%at", "--", workbook_path.name],
            cwd=workbook_path.parent,
        ).decode("utf-8").strip()
        if git_epoch_str:
            epoch_time = float(git_epoch_str)
            try:
                est_tz = zoneinfo.ZoneInfo("America/New_York")
                dt = datetime.fromtimestamp(epoch_time, est_tz)
            except Exception:
                dt = datetime.fromtimestamp(epoch_time)
            return format_custom_datetime(dt)
    except Exception:
        pass
    
    if workbook_path.exists():
        try:
            est_tz = zoneinfo.ZoneInfo("America/New_York")
            dt = datetime.fromtimestamp(workbook_path.stat().st_mtime, est_tz)
        except Exception:
            dt = datetime.fromtimestamp(workbook_path.stat().st_mtime)
        return format_custom_datetime(dt)
    return "Unavailable"


def format_week_date_range(year: int, week_number: int) -> str:
    start = date.fromisocalendar(year, week_number, 1)
    end = date.fromisocalendar(year, week_number, 7)
    if start.year == end.year and start.month == end.month:
        return f"{start.strftime('%B')} {start.day} to {end.strftime('%B')} {end.day}, {end.year}"
    return f"{start.strftime('%B')} {start.day}, {start.year} to {end.strftime('%B')} {end.day}, {end.year}"


def latest_dataset_text(rankings: pd.DataFrame, workbook_path: Path, modified_ns: int) -> Tuple[str, str]:
    latest_sort = rankings["Week Sort"].max()
    latest_row = rankings[rankings["Week Sort"] == latest_sort].iloc[0]
    week_number = int(latest_row["Week Number"])
    ranking_year = int(latest_row["Ranking Year"])
    range_text = format_week_date_range(ranking_year, week_number)
    return (
        f"Latest dataset: Week {week_number}, {ranking_year} ({range_text})",
        f"Data updated: {get_data_updated_text(workbook_path, modified_ns)}",
    )


def parse_week_sheet_name(sheet_name: str) -> Tuple[int, int, str]:
    match = WEEK_SHEET_RE.match(str(sheet_name).strip())
    if not match:
        raise ValueError(f"Not a weekly ranking sheet: {sheet_name}")
    week_number = int(match.group(1))
    ranking_year = int(match.group(2))
    return week_number, ranking_year, f"Week {week_number:02d}-{ranking_year}"


def clean_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def extract_hyperlink_formula_url(value):
    if not isinstance(value, str):
        return None
    formula = value.strip()
    if not formula.upper().startswith("=HYPERLINK("):
        return None
    match = re.search(r'=HYPERLINK\(\s*["\']([^"\']+)["\']', formula, flags=re.IGNORECASE)
    return match.group(1).strip() if match else None


def normalize_external_url(value):
    if not isinstance(value, str):
        return None
    url = value.strip()
    return url if url.lower().startswith(("http://", "https://")) else None


def coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    numeric_cols = ["Rank", "Year of Birth", "Points", "Tournaments", "Singles WTN", "Career Matches", "Singles W/L Career %", "Singles W/L-YTD %"]
    for col in numeric_cols:
        if col in df.columns:
            normalized = df[col].astype("string").str.strip().replace({"unavailable": pd.NA, "Unavailable": pd.NA, "UNAVAILABLE": pd.NA})
            df[col] = pd.to_numeric(normalized, errors="coerce")
    return df


def extract_profile_urls(ws, df: pd.DataFrame) -> List[object]:
    header_lookup = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value is not None}
    profile_col = header_lookup.get("Profile")
    if profile_col is None:
        return [pd.NA] * len(df)
    urls: List[object] = []
    for row_index in df.index:
        cell = ws.cell(row=int(row_index) + 2, column=profile_col)
        candidates = []
        if cell.hyperlink is not None:
            candidates.extend([cell.hyperlink.target, cell.hyperlink.location])
        candidates.extend([extract_hyperlink_formula_url(cell.value), cell.value])
        for candidate in candidates:
            external_url = normalize_external_url(candidate)
            if external_url:
                urls.append(external_url)
                break
        else:
            urls.append(pd.NA)
    return urls


def get_latest_profile_url(player_df: pd.DataFrame):
    if "Profile URL" not in player_df.columns:
        return None
    for url in player_df.sort_values("Week Sort", ascending=False)["Profile URL"].tolist():
        external_url = normalize_external_url(url)
        if external_url:
            return external_url
    return None


def normalize_avatar_url(url: str | None) -> str | None:
    if not isinstance(url, str):
        return None
    cleaned = url.strip()
    if not cleaned:
        return None
    cleaned = re.sub(r"\?.*$", "", cleaned)
    if cleaned.startswith("//"):
        cleaned = f"https:{cleaned}"
    return cleaned if cleaned.lower().startswith(("http://", "https://")) else None


def extract_avatar_url_from_html(page_html: str | None) -> str | None:
    if not isinstance(page_html, str) or not page_html.strip():
        return None

    for pattern in [
        r'<div[^>]*class=["\'][^"\']*profile-icon__img-inner[^"\']*["\'][^>]*style=["\'][^"\']*background-image\s*:\s*url\((?P<quote>["\']?)(?P<url>[^)]+?)(?P=quote)\)[^"\']*["\']',
        r'<img[^>]*data-testid=["\']profile\.header\.avatar\.image["\'][^>]*src=["\'](?P<url>https?://[^"\']+)["\']',
        r'<img[^>]*src=["\'](?P<url>https?://prod-cdn\.utrsports\.net/[^"\']+)["\']',
    ]:
        match = re.search(pattern, page_html, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return normalize_avatar_url(match.group("url"))
    return None


def player_initials(player_name: str) -> str:
    if not isinstance(player_name, str):
        return "?"
    parts = [part for part in re.split(r"\s+", player_name.strip()) if part.strip()]
    if not parts:
        return "?"
    letters = []
    for part in parts:
        letters.append(re.sub(r"[^A-Za-zÀ-ÖØ-öø-ÿ]", "", part)[:1].upper())
    if len(letters) <= 2:
        initial_text = "".join(letters[:2])
    else:
        initial_text = f"{letters[0]}{letters[-1]}"
    return initial_text[:2] or "?"


@st.cache_data(show_spinner=False)
def fetch_html_page(url: str | None) -> str | None:
    if not url:
        return None
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
        method="GET",
    )
    try:
        with urlopen(request, timeout=15) as response:
            return response.read().decode("utf-8", errors="replace")
    except (HTTPError, URLError, TimeoutError, OSError):
        return None


def resolve_player_avatar_url(selected_player: str, player_df: pd.DataFrame) -> str | None:
    profile_url = get_latest_profile_url(player_df)
    if profile_url:
        wtn_html = fetch_html_page(profile_url)
        avatar_url = extract_avatar_url_from_html(wtn_html)
        if avatar_url:
            return avatar_url

    if not UTR_LOOKUP_WORKBOOK.exists():
        return None
    try:
        profile_ids = load_utr_lookup(
            UTR_LOOKUP_WORKBOOK,
            UTR_LOOKUP_WORKBOOK.stat().st_mtime_ns,
        )
    except (OSError, ValueError, pd.errors.EmptyDataError):
        return None

    profile_id = profile_ids.get(selected_player)
    if profile_id is None:
        return None
    utr_url = f"https://app.utrsports.net/profiles/{profile_id}"
    utr_html = fetch_html_page(utr_url)
    return extract_avatar_url_from_html(utr_html)


def build_avatar_html(player_name: str, player_df: pd.DataFrame, class_name: str) -> str:
    """Build avatar markup, falling back to colour-seeded initials when no image resolves."""
    avatar_url = resolve_player_avatar_url(player_name, player_df)
    if avatar_url:
        return (
            f'<div class="{class_name}">'
            f'<img src="{escape(avatar_url, quote=True)}" alt="{escape(player_name, quote=True)} avatar" />'
            f"</div>"
        )
    palette = [
        ("#dfe7ee", "#b7c5cf"),
        ("#dce9f5", "#aec7dc"),
        ("#e7e3d7", "#c5c2b7"),
        ("#e4dce9", "#c7b5d1"),
    ]
    start_color, end_color = palette[sum(ord(character) for character in player_name) % len(palette)]
    return (
        f'<div class="{class_name} {class_name}--initials" '
        f'style="background: linear-gradient(135deg, {start_color} 0%, {end_color} 100%);">'
        f"{escape(player_initials(player_name))}</div>"
    )


@st.cache_data(show_spinner=False)
def load_rankings(excel_file: Path, modified_ns: int) -> pd.DataFrame:
    workbook = load_workbook(excel_file, data_only=False, read_only=False)
    sheet_names = [name for name in workbook.sheetnames if WEEK_SHEET_RE.match(str(name).strip())]

    def load_single_sheet(sheet_name: str) -> pd.DataFrame:
        ws = workbook[sheet_name]
        sheet_df = pd.DataFrame(ws.values)
        if sheet_df.empty:
            return pd.DataFrame()
        
        sheet_df.columns = sheet_df.iloc[0]
        sheet_df = sheet_df.iloc[1:].reset_index(drop=True)
        
        week_number, ranking_year, week_label = parse_week_sheet_name(sheet_name)
        df = clean_headers(sheet_df)
        if "Player" not in df.columns:
            return pd.DataFrame()
            
        df["Profile URL"] = extract_profile_urls(ws, df)
        
        df = df.dropna(how="all")
        df = df[df["Player"].notna()]
        df["Player"] = df["Player"].astype(str).str.strip()
        df = df[df["Player"] != ""]
        df["Week Number"] = week_number
        df["Ranking Year"] = ranking_year
        df["Week Label"] = week_label
        df["Week Sort"] = ranking_year * 100 + week_number
        if "Province" in df.columns:
            df["Province"] = df["Province"].apply(normalize_province_name)
        return df

    frames: List[pd.DataFrame] = []
    for sheet_name in sheet_names:
        df_sheet = load_single_sheet(sheet_name)
        if not df_sheet.empty:
            frames.append(df_sheet)

    if not frames:
        workbook.close()
        raise ValueError("No weekly sheets found. Expected sheet names like 'Week 27-2026'.")
    workbook.close()
    return coerce_numeric_columns(pd.concat(frames, ignore_index=True, sort=False)).sort_values(["Week Sort", "Rank", "Player"], na_position="last")


def format_value(metric: str, value) -> str:
    if pd.isna(value):
        return "unavailable"
    if metric in PERCENT_METRICS:
        return f"{value:.0%}"
    if metric == "Year of Birth":
        return f"{int(round(value))}"
    if metric in {"Rank", "Ontario Rank", "Tournaments", "Career Matches"}:
        return f"{int(round(value)):,}"
    if metric == "Singles WTN":
        return f"{value:.1f}"
    if metric == "Points":
        return f"{value:,.3f}"
    return f"{value:,.2f}"


def performance_change(metric: str, previous_value, current_value):
    if pd.isna(previous_value) or pd.isna(current_value):
        return pd.NA
    return previous_value - current_value if metric in LOWER_IS_BETTER else current_value - previous_value


def format_trend(metric: str, change) -> str:
    if pd.isna(change):
        return "\u2014"
    if abs(float(change)) < 1e-12:
        return "\u2192 0"
    arrow = "\u2191" if change > 0 else "\u2193"
    sign = "+" if change > 0 else ""
    if metric in PERCENT_METRICS:
        return f"{arrow} {sign}{change:.0%} pp"
    if metric in {"Rank", "Tournaments", "Career Matches"}:
        return f"{arrow} {sign}{change:.0f}"
    if metric == "Singles WTN":
        return f"{arrow} {sign}{change:.1f}"
    if metric == "Points":
        return f"{arrow} {sign}{change:,.3f}"
    return f"{arrow} {sign}{change:,.2f}"


def format_delta_for_metric_card(metric: str, change):
    if change is None or pd.isna(change):
        return None
    if abs(float(change)) < 1e-12:
        return "0"
    sign = "+" if change > 0 else ""
    if metric in PERCENT_METRICS:
        return f"{sign}{change:.0%} pp"
    if metric in {"Rank", "Tournaments", "Career Matches"}:
        return f"{sign}{change:.0f}"
    if metric == "Singles WTN":
        return f"{sign}{change:.1f}"
    if metric == "Points":
        return f"{sign}{change:,.3f}"
    return f"{sign}{change:,.2f}"


def value_with_week_trend(metric: str, previous_value, current_value) -> str:
    return f"{format_value(metric, current_value)} ({format_trend(metric, performance_change(metric, previous_value, current_value))})"


def color_trend_cell(value: object) -> str:
    text = str(value)
    if "\u2191" in text:
        return "color: #0f7b3f; font-weight: 600;"
    if "\u2193" in text:
        return "color: #b42318; font-weight: 600;"
    if "\u2192" in text:
        return "color: #6b7280;"
    return ""


def style_trend_table(display_df: pd.DataFrame, trend_cols: List[str]):
    cols = [c for c in trend_cols if c in display_df.columns]
    styler = display_df.style
    if not cols:
        return styler
    if hasattr(styler, "map"):
        return styler.map(color_trend_cell, subset=cols)
    return styler.applymap(color_trend_cell, subset=cols)


def prepare_linked_player_table(display_df: pd.DataFrame, rankings: pd.DataFrame) -> pd.DataFrame:
    """Attach source profile URLs so Streamlit can render player names as links."""
    if "Player" not in display_df.columns or "Profile URL" not in rankings.columns:
        return display_df
    profile_urls = (
        rankings.dropna(subset=["Profile URL"])
        .sort_values("Week Sort", ascending=False)
        .drop_duplicates("Player")
        .set_index("Player")["Profile URL"]
    )
    linked_df = display_df.copy()
    linked_df["Player"] = linked_df["Player"].map(
        lambda name: f"{profile_urls[name]}#{name}" if name in profile_urls.index else name
    )
    return linked_df


def render_player_dataframe(display_df: pd.DataFrame, rankings: pd.DataFrame, **kwargs) -> None:
    """Render dataframes with profile links for player-name cells where available."""
    source_df = display_df.data if hasattr(display_df, "data") else display_df
    linked_df = prepare_linked_player_table(source_df, rankings)
    column_config = {}
    if "Player" in linked_df.columns:
        column_config["Player"] = st.column_config.LinkColumn(
            "Player",
            display_text=r".*#(.*)$",
        )
    trend_columns = [
        column for column in linked_df.columns
        if column == "Rank Trend" or column in TOP_MOVER_METRIC_COLUMNS or column in LATEST_TRENDS_COLUMNS
    ]
    display = style_trend_table(linked_df, trend_columns) if trend_columns else linked_df
    st.dataframe(display, column_config=column_config, **kwargs)


def table_auto_height(row_count: int, row_height: int = 35, header_height: int = 42, padding: int = 8) -> int:
    return int(header_height + padding + max(row_count, 1) * row_height)


def get_latest_provincial_rank(rankings: pd.DataFrame, selected_player: str):
    latest_sort = rankings["Week Sort"].max()
    latest_df = rankings.loc[rankings["Week Sort"].eq(latest_sort)].copy()
    if "Province" not in latest_df.columns:
        return pd.NA
    latest_df["Province"] = latest_df["Province"].astype("string").str.strip()
    player_rows = latest_df.loc[latest_df["Player"].eq(selected_player)]
    if player_rows.empty or pd.isna(player_rows.iloc[0]["Province"]):
        return pd.NA
    ranked = latest_df.sort_values(["Province", "Rank", "Player"], na_position="last")
    ranked["Provincial Rank"] = ranked.groupby("Province", dropna=False).cumcount() + 1
    matches = ranked.loc[ranked["Player"].eq(selected_player), "Provincial Rank"]
    if matches.empty or pd.isna(matches.iloc[0]):
        return pd.NA
    return int(matches.iloc[0])


def render_wtn_donut_chart(wtn_value):
    if pd.isna(wtn_value):
        return "unavailable"
    
    val = float(wtn_value)
    fraction = max(0.0, min(1.0, (40.0 - val) / 39.0))
    
    filled_deg = fraction * 270
    rem_deg = 270 - filled_deg
    gap_deg = 90

    fig = go.Figure(go.Pie(
        values=[filled_deg, rem_deg, gap_deg],
        hole=0.75,
        direction="clockwise",
        rotation=225,
        sort=False,
        marker=dict(colors=["#1f77b4", "#e5e7eb", "rgba(0,0,0,0)"]),
        textinfo="none",
        hoverinfo="none"
    ))
    
    fig.update_layout(
        showlegend=False,
        autosize=True,
        margin=dict(l=20, r=20, t=20, b=20),
        annotations=[
            dict(
                text=f"<span style='font-family: inherit; font-size: 1.7rem; font-weight: 600; color: #111827;'>{val:.1f}</span>",
                x=0.5, y=0.5,
                showarrow=False,
                xref="paper",
                yref="paper"
            )
        ]
    )
    return fig


def build_player_history_display(player_df: pd.DataFrame) -> pd.DataFrame:
    chronological_df = player_df.sort_values("Week Sort").copy()
    excluded_cols = {"Club"}
    visible_cols = [c for c in DISPLAY_COLUMNS if c in chronological_df.columns and c not in excluded_cols]
    display_df = chronological_df[visible_cols + ["Week Sort"]].copy()
    for metric in NUMERIC_TREND_COLUMNS:
        if metric not in chronological_df.columns or metric not in display_df.columns:
            continue
        previous_values = chronological_df[metric].shift(1)
        display_df[metric] = [format_value(metric, v) if i == 0 else value_with_week_trend(metric, previous_values.iloc[i], v) for i, v in enumerate(chronological_df[metric].tolist())]
    display_df = display_df.sort_values("Week Sort", ascending=False).drop(columns=["Week Sort"])
    return display_df.rename(columns={"Week Label": "Week"}) if "Week Label" in display_df.columns else display_df


def build_latest_week_trends(rankings: pd.DataFrame) -> Tuple[pd.DataFrame, str, str]:
    week_sorts = sorted(rankings["Week Sort"].dropna().unique())
    if len(week_sorts) < 2:
        raise ValueError("At least two weekly sheets are needed to calculate latest week trends.")
    previous_sort, latest_sort = week_sorts[-2], week_sorts[-1]
    previous_label = rankings.loc[rankings["Week Sort"] == previous_sort, "Week Label"].iloc[0]
    latest_label = rankings.loc[rankings["Week Sort"] == latest_sort, "Week Label"].iloc[0]
    latest_df = rankings[rankings["Week Sort"] == latest_sort].copy().sort_values("Rank", na_position="last")
    previous_by_player = rankings[rankings["Week Sort"] == previous_sort].drop_duplicates("Player").set_index("Player")
    base_cols = [c for c in ["Rank", "Player", "Year of Birth", "Province", "Club"] if c in latest_df.columns]
    display_df = latest_df[base_cols].copy()
    if "Year of Birth" in display_df.columns:
        display_df["Year of Birth"] = display_df["Year of Birth"].apply(lambda v: format_value("Year of Birth", v))
    for metric in LATEST_TRENDS_COLUMNS:
        if metric not in latest_df.columns:
            continue
        previous_values = previous_by_player[metric].reindex(latest_df["Player"]).tolist() if metric in previous_by_player.columns else [pd.NA] * len(latest_df)
        display_df[metric] = [
            value_with_week_trend(metric, previous_value, current_value)
            for previous_value, current_value in zip(previous_values, latest_df[metric].tolist())
        ]
    return display_df, previous_label, latest_label


def build_top_movers(rankings: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, str, str]:
    week_sorts = sorted(rankings["Week Sort"].dropna().unique())
    if len(week_sorts) < 2:
        raise ValueError("At least two weekly sheets are needed to calculate top movers.")
    previous_sort, latest_sort = week_sorts[-2], week_sorts[-1]
    previous_label = rankings.loc[rankings["Week Sort"] == previous_sort, "Week Label"].iloc[0]
    latest_label = rankings.loc[rankings["Week Sort"] == latest_sort, "Week Label"].iloc[0]
    latest_df = rankings[rankings["Week Sort"] == latest_sort].drop_duplicates("Player").copy()
    previous_df = rankings[rankings["Week Sort"] == previous_sort].drop_duplicates("Player").copy()
    keep_cols = ["Player", "Rank"] + [c for c in ["Province"] + TOP_MOVER_METRIC_COLUMNS if c in rankings.columns]
    merged = latest_df[keep_cols].merge(previous_df[keep_cols], on="Player", suffixes=(" Latest", " Previous"), how="inner")
    merged["Rank Move"] = merged["Rank Previous"] - merged["Rank Latest"]
    merged = merged.dropna(subset=["Rank Move"])
    merged = merged[merged["Rank Move"] != 0]
    def make_table(source: pd.DataFrame, ascending: bool) -> pd.DataFrame:
        if source.empty:
            return pd.DataFrame(columns=["Player", "Province", "Previous Rank", "Current Rank", "Rank Trend"])
        source = source.sort_values("Rank Move", ascending=ascending).head(10)
        rows = []
        for _, row in source.iterrows():
            out = {"Player": row["Player"], "Province": row.get("Province Latest", "unavailable"), "Previous Rank": row["Rank Previous"], "Current Rank": row["Rank Latest"], "Rank Trend": format_trend("Rank", row["Rank Move"])}
            for metric in TOP_MOVER_METRIC_COLUMNS:
                latest_col, previous_col = f"{metric} Latest", f"{metric} Previous"
                if latest_col in row.index and previous_col in row.index:
                    out[metric] = value_with_week_trend(metric, row[previous_col], row[latest_col])
            rows.append(out)
        return pd.DataFrame(rows)
    return make_table(merged[merged["Rank Move"] > 0], ascending=False), make_table(merged[merged["Rank Move"] < 0], ascending=True), previous_label, latest_label


def calculate_ytd_trends(rankings: pd.DataFrame, player_df: pd.DataFrame) -> Tuple[pd.DataFrame, str, str]:
    oldest_sort, latest_sort = rankings["Week Sort"].min(), rankings["Week Sort"].max()
    oldest_label = rankings.loc[rankings["Week Sort"] == oldest_sort, "Week Label"].iloc[0]
    latest_label = rankings.loc[rankings["Week Sort"] == latest_sort, "Week Label"].iloc[0]
    start_df = player_df[player_df["Week Sort"] == oldest_sort].sort_values("Rank")
    end_df = player_df[player_df["Week Sort"] == latest_sort].sort_values("Rank")
    if start_df.empty or end_df.empty:
        missing = []
        if start_df.empty: missing.append(oldest_label)
        if end_df.empty: missing.append(latest_label)
        raise ValueError("Selected player is missing from the workbook comparison week(s): " + ", ".join(missing))
    start, end = start_df.iloc[0], end_df.iloc[0]
    rows = []
    for metric in YTD_TREND_METRICS:
        if metric not in player_df.columns:
            continue
        change = performance_change(metric, start.get(metric), end.get(metric))
        rows.append({"Metric": metric, f"Oldest ({oldest_label})": format_value(metric, start.get(metric)), f"Latest ({latest_label})": format_value(metric, end.get(metric)), "Trend": format_trend(metric, change), "Metric card delta": format_delta_for_metric_card(metric, change)})
    return pd.DataFrame(rows), oldest_label, latest_label


def get_latest_top_n(rankings: pd.DataFrame, n: int) -> Tuple[pd.DataFrame, str]:
    latest_sort = rankings["Week Sort"].max()
    latest_label = rankings.loc[rankings["Week Sort"] == latest_sort, "Week Label"].iloc[0]
    return rankings[rankings["Week Sort"] == latest_sort].copy().sort_values("Rank", na_position="last").head(n), latest_label


def style_distribution_chart(fig):
    fig.update_traces(marker_line_color="rgba(31, 41, 55, 0.95)", marker_line_width=1, texttemplate="%{y}", textposition="outside", cliponaxis=False)
    fig.update_layout(title_x=0.5, uniformtext_minsize=10, uniformtext_mode="show")
    return fig


def chart_multi_player_metric(rankings: pd.DataFrame, players: List[str], metric: str):
    if not players:
        st.info("Please select at least one player to chart.")
        return
    chart_df = rankings[["Week Label", "Week Sort", "Player", metric]].dropna(subset=[metric]).sort_values("Week Sort")
    if chart_df.empty:
        st.info(f"No numeric data available for {metric} for the selected players.")
        return
    fig = px.line(chart_df, x="Week Label", y=metric, color="Player", markers=True, title=f"{metric} Comparison by Week", labels={"Week Label": "Week", metric: metric})
    fig.update_layout(xaxis_title="Week", yaxis_title=metric, hovermode="x unified", title_x=0.5)
    if metric in LOWER_IS_BETTER:
        fig.update_yaxes(autorange="reversed")
    if "%" in metric:
        fig.update_yaxes(tickformat=".0%")
    st.plotly_chart(fig, width="stretch")


def build_player_metric_figure(player_df: pd.DataFrame, metric: str):
    """Build a report-ready weekly player chart for a numeric metric."""
    chart_df = player_df[["Week Label", "Week Sort", metric]].dropna(subset=[metric]).sort_values("Week Sort")
    if chart_df.empty:
        return None
    figure = px.line(chart_df, x="Week Label", y=metric, markers=True, title=f"{metric} by week")
    figure.update_layout(
        title_x=0.5,
        margin=dict(l=30, r=20, t=50, b=35),
        height=260,
        xaxis_title="Week",
        yaxis_title=metric,
    )
    if metric in LOWER_IS_BETTER:
        figure.update_yaxes(autorange="reversed")
    if "%" in metric:
        figure.update_yaxes(tickformat=".0%")
    return figure


def convert_markdown_to_reportlab(text: str) -> str:
    """Convert markdown formatting to ReportLab-compatible HTML tags.
    
    Supports:
    - **text** or __text__ → <b>text</b> (bold)
    - *text* or _text_ → <i>text</i> (italic)
    - - item → • item (bullet points)
    """
    # Escape HTML first to avoid double-escaping
    text = escape(text)
    
    # Convert bold: **text** or __text__
    text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
    text = re.sub(r'__(.+?)__', r'<b>\1</b>', text)
    
    # Convert italic: *text* or _text_ (but not in **text** or __text__)
    text = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'<i>\1</i>', text)
    text = re.sub(r'(?<!_)_(?!_)(.+?)(?<!_)_(?!_)', r'<i>\1</i>', text)
    
    # Convert bullet points: - item → • item
    if text.startswith('- '):
        text = '• ' + text[2:]
    
    return text


def build_pdf_line_chart(player_df: pd.DataFrame, metric: str) -> Drawing | None:
    """Create a theme-independent vector line chart for a PDF report."""
    chart_df = player_df[["Week Label", "Week Sort", metric]].dropna(subset=[metric]).sort_values("Week Sort")
    if chart_df.empty:
        return None

    width, height = 6.7 * inch, 2.45 * inch
    left, right, top, bottom = 0.55 * inch, 0.2 * inch, 0.32 * inch, 0.45 * inch
    plot_width, plot_height = width - left - right, height - top - bottom
    values = chart_df[metric].astype(float).tolist()
    low, high = min(values), max(values)
    if abs(high - low) < 1e-12:
        padding = max(abs(high) * 0.05, 1.0)
        low, high = low - padding, high + padding
    else:
        padding = (high - low) * 0.08
        low, high = low - padding, high + padding

    drawing = Drawing(width, height)
    drawing.add(Rect(left, bottom, plot_width, plot_height, strokeColor=colors.HexColor("#D1D5DB"), fillColor=colors.white))
    drawing.add(String(left, height - 0.14 * inch, metric, fontName="Helvetica-Bold", fontSize=9, fillColor=colors.HexColor("#1F2937")))
    
    # Add x-axis labels: first, last, and intermediate points (every Nth data point)
    num_labels = min(5, len(chart_df))  # Target 5 labels, but at most the number of data points
    if num_labels > 2:
        label_indices = sorted(set([0, len(chart_df) - 1] + 
                                   [int(i * (len(chart_df) - 1) / (num_labels - 1)) for i in range(1, num_labels - 1)]))
    else:
        label_indices = [0] if len(chart_df) > 0 else []
        if len(chart_df) > 1:
            label_indices.append(len(chart_df) - 1)
    
    for idx in label_indices:
        if idx < len(chart_df):
            label_text = chart_df["Week Label"].iloc[idx]
            # Calculate x position proportionally
            count = max(len(values) - 1, 1)
            x_pos = left + idx / count * plot_width
            # Center the text around the calculated position
            drawing.add(String(x_pos - 0.3 * inch, 0.08 * inch, label_text, fontSize=6.5, fillColor=colors.HexColor("#6B7280")))
    
    # Y-axis labels: swap for LOWER_IS_BETTER metrics since scale is inverted
    if metric in LOWER_IS_BETTER:
        drawing.add(String(0.03 * inch, bottom + plot_height - 2, format_value(metric, low), fontSize=6.5, fillColor=colors.HexColor("#6B7280")))
        drawing.add(String(0.03 * inch, bottom - 2, format_value(metric, high), fontSize=6.5, fillColor=colors.HexColor("#6B7280")))
    else:
        drawing.add(String(0.03 * inch, bottom + plot_height - 2, format_value(metric, high), fontSize=6.5, fillColor=colors.HexColor("#6B7280")))
        drawing.add(String(0.03 * inch, bottom - 2, format_value(metric, low), fontSize=6.5, fillColor=colors.HexColor("#6B7280")))

    points = []
    count = max(len(values) - 1, 1)
    for index, value in enumerate(values):
        x = left + index / count * plot_width
        scaled = (value - low) / (high - low)
        if metric in LOWER_IS_BETTER:
            scaled = 1 - scaled
        y = bottom + scaled * plot_height
        points.extend([x, y])
    drawing.add(PolyLine(points, strokeColor=colors.HexColor("#1F77B4"), strokeWidth=1.5))
    
    # Add data labels and point markers
    for index in range(0, len(points), 2):
        x_coord = points[index]
        y_coord = points[index + 1]
        # Add point marker
        drawing.add(Rect(x_coord - 1.5, y_coord - 1.5, 3, 3, strokeColor=colors.HexColor("#1F77B4"), fillColor=colors.HexColor("#1F77B4")))
        # Add data label above the point
        data_point_index = index // 2
        if data_point_index < len(values):
            label = format_value(metric, values[data_point_index])
            # Position label above point with better visibility
            # Use larger font and position relative to point with bounds checking
            label_y = y_coord + 0.15 * inch
            # Center-align by offsetting left by approximate text width
            label_x = x_coord - len(label) * 1.5
            drawing.add(String(label_x, label_y, label, fontSize=7, fontName="Helvetica-Bold", fillColor=colors.HexColor("#1F77B4")))
    return drawing


def format_top_players_table(source_df: pd.DataFrame, include_ontario_rank: bool = False) -> pd.DataFrame:
    player_cols = ["Rank", "Player", "Year of Birth", "Points", "Tournaments", "Province", "Club", "Singles WTN", "Career Matches", "Singles W/L-Career", "Singles W/L-YTD", "Singles W/L Career %", "Singles W/L-YTD %"]
    player_cols = [col for col in player_cols if col in source_df.columns]
    sorted_source = source_df.sort_values("Rank", na_position="last").copy()
    if include_ontario_rank:
        sorted_source.insert(sorted_source.columns.get_loc("Rank") + 1 if "Rank" in sorted_source.columns else 0, "Ontario Rank", range(1, len(sorted_source) + 1))
        if "Ontario Rank" not in player_cols:
            player_cols.insert(1 if "Rank" in player_cols else 0, "Ontario Rank")
        if "Province" in player_cols:
            player_cols.remove("Province")
    top_players = sorted_source[player_cols].copy()
    for metric in ["Rank", "Ontario Rank", "Year of Birth", "Points", "Tournaments", "Singles WTN", "Career Matches", "Singles W/L Career %", "Singles W/L-YTD %"]:
        if metric in top_players.columns:
            top_players[metric] = top_players[metric].apply(lambda value, metric=metric: format_value(metric, value))
    return top_players


def render_top_players_table(source_df: pd.DataFrame, title: str, height: int | None, include_ontario_rank: bool = False):
    st.subheader(title, anchor=False)
    display_df = format_top_players_table(source_df, include_ontario_rank=include_ontario_rank)
    render_player_dataframe(display_df, source_df, width="stretch", height=height or table_auto_height(len(display_df)), hide_index=True)


def centered_number_table(df: pd.DataFrame):
    styler = df.style
    if "Number of players" in df.columns:
        styler = styler.set_properties(subset=["Number of players"], **{"text-align": "center"})
    return styler


def render_top_n_analytics(topn_df: pd.DataFrame, latest_label: str, n: int, *, show_province_chart: bool = True, clubs_include_province: bool = True, include_players_table: bool = False, players_table_title: str | None = None, players_table_height: int | None = 760, caption_prefix: str | None = None, include_ontario_rank: bool = False):
    caption = caption_prefix or f"Top {n} analysis"
    st.caption(f"{caption} based on {latest_label} data.")
    yob_counts = topn_df["Year of Birth"].dropna().astype(int).astype(str).value_counts().sort_index().reset_index()
    yob_counts.columns = ["Year of Birth", "Count"]
    club_source = topn_df.copy()
    club_source["Club"] = club_source["Club"].fillna("Unavailable").astype(str).str.strip().replace("", "Unavailable")
    club_source["Province"] = club_source["Province"].fillna("Unavailable").astype(str).str.strip().replace("", "Unavailable") if "Province" in club_source.columns else "Unavailable"
    if clubs_include_province:
        club_counts = (club_source.groupby("Club", as_index=False).agg(Province=("Province", lambda values: ", ".join(sorted(set(values)))), **{"Number of players": ("Player", "nunique")}).sort_values(["Number of players", "Club"], ascending=[False, True]).head(10))
    else:
        club_counts = (club_source.groupby("Club", as_index=False).agg(**{"Number of players": ("Player", "nunique")}).sort_values(["Number of players", "Club"], ascending=[False, True]).head(10))

    if show_province_chart:
        province_counts = topn_df["Province"].fillna("Unavailable").value_counts().reset_index()
        province_counts.columns = ["Province", "Count"]
        pie_col1, pie_col2 = st.columns(2)
        with pie_col1:
            fig_province = px.pie(province_counts, names="Province", values="Count", title=f"Top {n} by Province")
            fig_province.update_layout(title_x=0.5)
            st.plotly_chart(fig_province, width="stretch")
        with pie_col2:
            fig_yob = px.pie(yob_counts, names="Year of Birth", values="Count", title=f"Top {n} by Year of Birth")
            fig_yob.update_layout(title_x=0.5)
            st.plotly_chart(fig_yob, width="stretch")
    else:
        row_col1, row_col2 = st.columns(2)
        with row_col1:
            fig_yob = px.pie(yob_counts, names="Year of Birth", values="Count", title=f"Top {n} by Year of Birth")
            fig_yob.update_layout(title_x=0.5)
            st.plotly_chart(fig_yob, width="stretch")
        with row_col2:
            table_left, table_mid, table_right = st.columns([1, 8, 1])
            with table_mid:
                st.subheader(f"Top 10 clubs by number of Top {n} players", anchor=False)
                st.dataframe(centered_number_table(club_counts), width="stretch", hide_index=True)
        st.markdown("<div style='height: 2.25rem;'></div>", unsafe_allow_html=True)

    avg_wtn = topn_df["Singles WTN"].dropna().mean() if "Singles WTN" in topn_df.columns else pd.NA
    avg_matches = topn_df["Career Matches"].dropna().mean() if "Career Matches" in topn_df.columns else pd.NA
    avg_ytd = topn_df["Singles W/L-YTD %"].dropna().mean() if "Singles W/L-YTD %" in topn_df.columns else pd.NA
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("Average WTN", "unavailable" if pd.isna(avg_wtn) else f"{avg_wtn:.1f}")
    metric_col2.metric("Average number of career matches", "unavailable" if pd.isna(avg_matches) else f"{avg_matches:.0f}")
    metric_col3.metric("Average Singles W/L YTD", "unavailable" if pd.isna(avg_ytd) else f"{avg_ytd:.0%}")
    hist_col1, hist_col2 = st.columns(2)
    with hist_col1:
        wtn_data = topn_df[["Singles WTN"]].dropna() if "Singles WTN" in topn_df.columns else pd.DataFrame(columns=["Singles WTN"])
        if wtn_data.empty:
            st.info(f"No WTN data available for the Top {n} players.")
        else:
            fig_wtn = px.histogram(wtn_data, x="Singles WTN", title="WTN Distribution", labels={"Singles WTN": "Singles WTN"})
            fig_wtn.update_traces(xbins=dict(size=0.5))
            fig_wtn.update_layout(title_x=0.5, yaxis_title="Number of players")
            st.plotly_chart(style_distribution_chart(fig_wtn), width="stretch")
    with hist_col2:
        match_data = topn_df[["Career Matches"]].dropna() if "Career Matches" in topn_df.columns else pd.DataFrame(columns=["Career Matches"])
        if match_data.empty:
            st.info(f"No career matches data available for the Top {n} players.")
        else:
            fig_matches = px.histogram(match_data, x="Career Matches", title="Career Matches Distribution", labels={"Career Matches": "Career Matches"})
            fig_matches.update_traces(xbins=dict(size=25))
            fig_matches.update_layout(title_x=0.5, yaxis_title="Number of players")
            st.plotly_chart(style_distribution_chart(fig_matches), width="stretch")
    if show_province_chart:
        st.subheader(f"Top 10 clubs by number of Top {n} players", anchor=False)
        st.dataframe(centered_number_table(club_counts), width="stretch", hide_index=True)
    if include_players_table:
        render_top_players_table(topn_df, players_table_title or f"Top {n} players by rank", players_table_height, include_ontario_rank=include_ontario_rank)


def render_top_n_tab(rankings: pd.DataFrame, n: int):
    topn_df, latest_label = get_latest_top_n(rankings, n)
    render_top_n_analytics(topn_df, latest_label, n, include_players_table=(n == 20), players_table_title="Top 20 players by rank" if n == 20 else None, players_table_height=760)


def render_top_50_ontario_tab(rankings: pd.DataFrame):
    latest_sort = rankings["Week Sort"].max()
    latest_label = rankings.loc[rankings["Week Sort"] == latest_sort, "Week Label"].iloc[0]
    latest_df = rankings[rankings["Week Sort"] == latest_sort].copy()
    ontario_df = latest_df[latest_df["Province"].astype(str).str.strip().eq("Ontario")]
    top50_ontario_df = ontario_df.sort_values("Rank", na_position="last").head(50)
    if top50_ontario_df.empty:
        st.info("No Ontario players were found in the latest week's data.")
        return
    render_top_n_analytics(top50_ontario_df, latest_label, 50, show_province_chart=False, clubs_include_province=False, include_players_table=True, players_table_title="Top 50 Ontario players by rank", players_table_height=None, caption_prefix="Top 50 Ontario analysis", include_ontario_rank=True)


def render_multi_player_download_tab(rankings: pd.DataFrame):
    st.subheader("Multi-Player Raw Data Download", anchor=False)
    st.caption("Paste player names below (each name on a new line) to retrieve their data from the most recently available week and download it as an Excel spreadsheet.")

    names_input = st.text_area(
        "Player Names",
        placeholder="Ela Velic\nPlayer Two\nPlayer Three",
        height=150
    )

    if st.button("Retrieve Data", type="primary"):
        if not names_input.strip():
            st.warning("Please enter at least one player name.")
            return

        target_players_raw = [name.strip() for name in names_input.splitlines() if name.strip()]
        
        rankings = rankings.copy()
        rankings["Player_Lower"] = rankings["Player"].str.lower()
        target_players_lower = [name.lower() for name in target_players_raw]

        filtered_df = rankings[rankings["Player_Lower"].isin(target_players_lower)].copy()

        if filtered_df.empty:
            st.error("No data found for any of the specified player names. Please check the spelling and try again.")
            return

        idx_latest = filtered_df.groupby("Player_Lower")["Week Sort"].idxmax()
        filtered_df = filtered_df.loc[idx_latest].drop(columns=["Player_Lower"])

        cols_to_drop = ["Profile", "Profile URL", "Week Number", "Ranking Year", "Week Sort"]
        filtered_df = filtered_df.drop(columns=[c for c in cols_to_drop if c in filtered_df.columns])

        sort_cols = [c for c in ["Player", "Week Label"] if c in filtered_df.columns]
        if sort_cols:
            filtered_df = filtered_df.sort_values(sort_cols)

        st.success(f"Retrieved the most recent week's record for {filtered_df['Player'].nunique()} player(s).")
        render_player_dataframe(filtered_df, rankings, width="stretch", hide_index=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            filtered_df.to_excel(writer, index=False, sheet_name="Player Data")
        excel_binary_data = output.getvalue()

        st.download_button(
            label="Download Data as Excel (.xlsx)",
            data=excel_binary_data,
            file_name="multi_player_latest_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def render_ai_analysis_tab(selected_player: str, player_df: pd.DataFrame, rankings: pd.DataFrame):
    st.subheader(f"AI Analysis for {selected_player}", anchor=False)
    st.caption("Summarizing the player's history, charts, and YTD trends using Google Gemini.")

    if "GEMINI_API_KEY" not in st.secrets:
        st.warning("Google Gemini API key not found. Please add `GEMINI_API_KEY` to your Streamlit secrets.")
        return

    if st.button("Generate AI Analysis", type="primary"):
        try:
            player_history = build_player_history_display(player_df).to_string()
        except Exception:
            player_history = player_df.to_string()

        try:
            ytd_df, _, _ = calculate_ytd_trends(rankings, player_df)
            ytd_summary = ytd_df.to_string()
        except Exception:
            ytd_summary = "YTD trends data unavailable."

        prompt = f"""
        You are an expert tennis analyst. Analyze the following statistics for the youth tennis player '{selected_player}':

        1. Player Match & Ranking History:
        {player_history}

        2. Year-to-Date (YTD) Trends:
        {ytd_summary}

        Please provide a concise, professional performance summary structured cleanly under these exact markdown headings:
        ### Overall Performance Overview
        ### Key Strengths & Progress
        ### Areas for Improvement
        ### Outlook & Recommendations
        """

        with st.spinner("Analyzing player performance with Google Gemini..."):
            try:
                client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])
                response = client.models.generate_content(
                    model="gemini-3.6-flash",
                    contents=prompt,
                )
                st.markdown(response.text)
            except Exception as e:
                st.error(f"Failed to generate AI analysis: {e}")


def render_sidebar_summary(weeks_loaded: int, players_found: int, ranking_rows: int):
    st.markdown("""
        <style>
        section[data-testid="stSidebar"] > div:first-child { padding-bottom: 12rem; }
        .sidebar-summary-fixed { position: fixed; bottom: 0.65rem; left: 0.75rem; width: 250px; max-width: calc(100vw - 1.5rem); padding: 0.65rem 0.75rem; border: none; box-shadow: none; border-radius: 0.5rem; background: var(--secondary-background-color); z-index: 1000; }
        .sidebar-summary-title { font-size: 0.95rem; font-weight: 700; margin-bottom: 0.5rem; }
        .sidebar-summary-row { display: flex; justify-content: space-between; gap: 0.6rem; margin: 0.22rem 0; font-size: 0.82rem; }
        .sidebar-summary-value { font-weight: 700; }
        </style>
        """, unsafe_allow_html=True)
    st.sidebar.markdown(f"""
        <div class="sidebar-summary-fixed">
            <div class="sidebar-summary-title">Database Summary</div>
            <div class="sidebar-summary-row"><span>Number of weeks of Data</span><span class="sidebar-summary-value">{weeks_loaded:,}</span></div>
            <div class="sidebar-summary-row"><span>Total Players</span><span class="sidebar-summary-value">{players_found:,}</span></div>
            <div class="sidebar-summary-row"><span>Total Rows of Data</span><span class="sidebar-summary-value">{ranking_rows:,}</span></div>
        </div>
        """, unsafe_allow_html=True)


def parse_win_loss_ytd(val):
    if pd.isna(val):
        return None, None
    s = str(val).strip()
    if not s or s.lower() in {"unavailable", "nan", "none"}:
        return None, None
    parts = s.split("/")
    if len(parts) != 2:
        return None, None
    try:
        w = int(float(parts[0].strip()))
        l = int(float(parts[1].strip()))
        return w, l
    except ValueError:
        return None, None


import sys
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

backend = sys.modules[__name__]
st.set_page_config(page_title="OTA Girls U14 Ranking Dashboard", layout="wide")

VIEWS_BY_SECTION = {
    "Player": ["Player history", "Player Charts", "YTD Player trends", "AI Analysis"],
    "Rankings": ["Top movers", "Latest Week Trends"],
    "Leaderboards": ["Top 100", "Top 50", "Top 20", "Top 50 Ontario"],
        "Tools": ["Multi-Player Download", "Head-to-Head"],
}
UTR_LOOKUP_WORKBOOK = Path(__file__).with_name("list.xlsx")
UTR_PROFILE_URL = "https://api.utrsports.net/v1/player/{player_id}/profile"
H2H_BACKEND_URL = st.secrets.get("H2H_BACKEND_URL", "").rstrip("/")
H2H_BACKEND_API_KEY = st.secrets.get("H2H_BACKEND_API_KEY", "")
H2H_MATCH_HISTORY_URL = st.secrets.get(
    "H2H_MATCH_HISTORY_URL", f"{H2H_BACKEND_URL}/common-matches"
).rstrip("/")
H2H_COMMON_COMPETITORS_URL = st.secrets.get(
    "H2H_COMMON_COMPETITORS_URL", f"{H2H_BACKEND_URL}/common-competitors"
).rstrip("/")


def apply_styles() -> None:
    """Apply all custom presentation rules in one place."""
    st.markdown(
        """
        <style>
        section[data-testid="stSidebar"] > div:first-child { padding-bottom: 13rem; }
        .player-header {
            border-bottom: 1px solid #d7dce3;
            padding: 0.15rem 0 0.9rem;
            margin-bottom: 1rem;
        }
        .player-header__row {
            display: flex;
            align-items: center;
            gap: 1.2rem;
        }
        .player-header__avatar {
            width: 110px;
            height: 110px;
            flex-shrink: 0;
            border-radius: 50%;
            border: 6px solid #b8c0c9;
            background: linear-gradient(135deg, #c9ced6 0%, #b3bac5 100%);
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
            box-sizing: border-box;
            box-shadow: inset 0 0 0 1px rgba(67, 74, 86, 0.08);
        }
        .player-header__avatar img {
            width: 100%;
            height: 100%;
            object-fit: cover;
            display: block;
        }
        .player-header__avatar--initials {
            font-size: 2.2rem;
            font-weight: 800;
            letter-spacing: 0.06em;
            color: #2b2d30;
            background: linear-gradient(135deg, #dfe7ee 0%, #bec9d3 100%);
        }
        .player-header__name {
            font-size: 2rem;
            font-weight: 700;
            line-height: 1.2;
        }
        .player-header__meta {
            color: #5b6472;
            font-size: 0.95rem;
            margin-top: 0.35rem;
        }
        .player-metrics {
            display: grid;
            grid-template-columns: repeat(6, minmax(0, 1fr));
            gap: 1rem;
            padding: 0.15rem 0 1.2rem;
        }
        .player-metric {
            text-align: center;
        }
        .player-metric__label {
            color: var(--text-color);
            font-size: 0.92rem;
            font-weight: 600;
            line-height: 1.35;
            min-height: 1.35rem;
        }
        .player-metric__value {
            color: var(--text-color);
            font-size: 2.2rem;
            font-weight: 400;
            line-height: 1.2;
            margin-top: 0.1rem;
        }
        .player-metric__value-area {
            display: flex;
            flex-direction: column;
            justify-content: center;
            min-height: 2.74rem;
            margin-top: 0.1rem;
        }
        .player-metric--win-loss .player-metric__value {
            font-size: 1.8rem;
            line-height: 1.05;
            margin-top: 0;
        }
        .player-metric__win-loss-bar {
            display: flex;
            width: 120px;
            height: 6px;
            overflow: hidden;
            margin: 0.3rem auto 0;
            border-radius: 3px;
            background: #e5e7eb;
        }
        .player-metric__wins { background: #22c55e; }
        .player-metric__losses { background: #ef4444; }
        .player-metric--utr .player-metric__value-area,
        .player-metric--wtn .player-metric__value-area {
            flex-direction: row;
            align-items: center;
            gap: 0.35rem;
        }
        .player-metric__rating-donut,
        .player-metric__wtn-donut {
            flex: 0 0 2.55rem;
            width: 2.55rem;
            height: 2.55rem;
            border-radius: 50%;
            -webkit-mask: radial-gradient(farthest-side, transparent calc(100% - 0.5rem), #000 calc(100% - 0.48rem));
            mask: radial-gradient(farthest-side, transparent calc(100% - 0.5rem), #000 calc(100% - 0.48rem));
            background: conic-gradient(
                from 225deg,
                #1f77b4 0deg var(--wtn-filled),
                #e5e7eb var(--wtn-filled) 270deg,
                transparent 270deg 360deg
            );
        }
        .h2h-scoreband {
            border: 1px solid #d7dce3;
            border-radius: 0.9rem;
            background: linear-gradient(95deg, #f6f9fe 0%, #ffffff 50%, #fff7f7 100%);
            padding: 1.05rem 1.2rem;
            margin: 0.75rem 0 0.9rem;
            display: grid;
            grid-template-columns: 1fr auto 1fr;
            align-items: center;
            gap: 1rem;
        }
        .h2h-scoreband__name {
            font-size: 1.35rem;
            font-weight: 700;
            color: #1d4d8f;
            line-height: 1.25;
        }
        .h2h-scoreband__name a {
            color: inherit;
        }
        .h2h-scoreband__name--right {
            text-align: right;
            color: #d63b52;
        }
        .h2h-scoreband__player {
            display: flex;
            align-items: center;
            gap: 0.7rem;
        }
        .h2h-scoreband__player--right {
            flex-direction: row-reverse;
        }
        .h2h-avatar {
            width: 66px;
            height: 66px;
            flex-shrink: 0;
            border-radius: 50%;
            border: 3.6px solid #b8c0c9;
            background: linear-gradient(135deg, #c9ced6 0%, #b3bac5 100%);
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
            box-sizing: border-box;
            box-shadow: inset 0 0 0 1px rgba(67, 74, 86, 0.08);
        }
        .h2h-avatar img {
            width: 100%;
            height: 100%;
            object-fit: cover;
            display: block;
        }
        .h2h-avatar--initials {
            font-size: 1.32rem;
            font-weight: 800;
            letter-spacing: 0.06em;
            color: #2b2d30;
        }
        .h2h-scoreband__player .h2h-avatar {
            border-color: #1d4d8f;
        }
        .h2h-scoreband__player--right .h2h-avatar {
            border-color: #d63b52;
        }
        .h2h-scoreband__score {
            display: flex;
            align-items: flex-end;
            justify-content: center;
            gap: 0.75rem;
        }
        .h2h-scoreband__score-left,
        .h2h-scoreband__score-right {
            font-size: 2.8rem;
            font-weight: 700;
            line-height: 1;
        }
        .h2h-scoreband__score-left { color: #1d4d8f; }
        .h2h-scoreband__score-right { color: #d63b52; }
        .h2h-scoreband__vs {
            font-size: 1.2rem;
            font-weight: 700;
            color: #6b7280;
            padding-bottom: 0.4rem;
        }
        .h2h-scoreband__labels {
            display: flex;
            justify-content: center;
            gap: 2.2rem;
            margin-top: 0.15rem;
            font-size: 0.8rem;
            font-weight: 700;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: #6b7280;
        }
        .h2h-not-competed {
            margin-top: 0.35rem;
            text-align: center;
            font-size: 0.95rem;
            font-weight: 600;
            color: #4b5563;
        }
        .h2h-metrics-table {
            width: 100%;
            table-layout: fixed;
            border-collapse: collapse;
            border: 1px solid #d7dce3;
            border-radius: 0.75rem;
            overflow: hidden;
            background: #ffffff;
        }
        .h2h-metrics-table th,
        .h2h-metrics-table td {
            border: 1px solid #e5e7eb;
            padding: 0.7rem 0.8rem;
            text-align: center;
            vertical-align: middle;
            font-size: 0.98rem;
        }
        .h2h-metrics-table th {
            font-weight: 700;
            background: #f8fafc;
        }
        .h2h-metrics-table .player-header--left {
            color: #1d4d8f;
        }
        .h2h-metrics-table .player-header--right {
            color: #9f2238;
        }
        .h2h-metrics-table .metric-header {
            background: #ffffff;
        }
        .h2h-metrics-table .metric-label {
            font-weight: 700;
            background: #f3f4f6;
            white-space: nowrap;
        }
        .h2h-history {
            width: 100%;
            margin-top: 0;
            border-collapse: collapse;
            border: 1px solid #d7dce3;
            border-radius: 0.75rem;
            overflow: hidden;
            background: #ffffff;
            table-layout: fixed;
        }
        .h2h-history th,
        .h2h-history td {
            border: 1px solid #e5e7eb;
            padding: 0.7rem 0.75rem;
            text-align: center;
            vertical-align: middle;
            font-size: 0.92rem;
        }
        .h2h-history th {
            background: #f8fafc;
            color: #1d3155;
            font-weight: 700;
        }
        .h2h-history__winner {
            color: #d63b52;
            font-weight: 600;
        }
        .h2h-section-title {
            margin-top: 1.8rem;
            margin-bottom: 0.7rem;
            color: #1d3155;
            font-size: 1.2rem;
            font-weight: 700;
            text-align: center;
        }
        .h2h-history__empty {
            margin-top: 0;
            padding: 0.85rem;
            border: 1px solid #e5e7eb;
            border-radius: 0.75rem;
            text-align: center;
            color: #6b7280;
        }
        .h2h-common-competitors {
            width: 100%;
            margin-top: 0;
            border-collapse: collapse;
            border: 1px solid #d7dce3;
            border-radius: 0.75rem;
            overflow: hidden;
            background: #ffffff;
            table-layout: fixed;
        }
        .h2h-common-competitors th,
        .h2h-common-competitors td {
            border: 1px solid #e5e7eb;
            padding: 0.65rem 0.55rem;
            text-align: center;
            vertical-align: middle;
            font-size: 0.9rem;
        }
        .h2h-common-competitors th {
            background: #f8fafc;
            color: #1d3155;
            font-weight: 700;
        }
        .h2h-common-competitors tfoot th,
        .h2h-common-competitors tfoot td {
            background: #f3f4f6;
            color: #1d3155;
            font-weight: 700;
        }
        .h2h-common-competitors__left {
            color: #1d4d8f;
        }
        .h2h-common-competitors__right {
            color: #9f2238;
        }
        .h2h-common-competitors__name {
            text-align: left;
        }
        .h2h-common-competitors__name sup {
            font-size: 0.68em;
            color: #6b7280;
            white-space: nowrap;
        }
        .h2h-common-competitors__name a {
            color: inherit;
            text-decoration: none;
        }
        .h2h-common-competitors__name a:hover {
            text-decoration: underline;
        }
        .sidebar-data-status {
            position: fixed;
            bottom: 0.75rem;
            left: 0.75rem;
            width: 250px;
            max-width: calc(100vw - 1.5rem);
            color: #6b7280;
            font-size: 0.76rem;
            line-height: 1.45;
            padding: 0.65rem 0.75rem;
            border-top: 1px solid #d7dce3;
            background: var(--secondary-background-color);
            z-index: 1000;
        }
        .sidebar-data-status__title {
            color: #4b5563;
            font-weight: 600;
            margin-bottom: 0.3rem;
        }
        @media (max-width: 640px) {
            .player-header__name { font-size: 1.55rem; }
            .player-header__meta { line-height: 1.55; }
            .sidebar-data-status { width: calc(100vw - 1.5rem); }
            .player-metrics { grid-template-columns: repeat(2, minmax(0, 1fr)); }
            .player-metric__value { font-size: 1.85rem; }
            .h2h-scoreband {
                grid-template-columns: 1fr;
                text-align: center;
                gap: 0.55rem;
            }
            .h2h-scoreband__name,
            .h2h-scoreband__name--right {
                text-align: center;
                font-size: 1.15rem;
            }
            .h2h-scoreband__player,
            .h2h-scoreband__player--right {
                justify-content: center;
            }
            .h2h-scoreband__score-left,
            .h2h-scoreband__score-right {
                font-size: 2.2rem;
            }
            .h2h-metrics-table th,
            .h2h-metrics-table td {
                font-size: 0.9rem;
                padding: 0.6rem;
            }
            .h2h-history th,
            .h2h-history td {
                font-size: 0.82rem;
                padding: 0.55rem 0.35rem;
            }
            .h2h-common-competitors th,
            .h2h-common-competitors td {
                font-size: 0.78rem;
                padding: 0.5rem 0.25rem;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


@st.cache_data(show_spinner=False)
def load_utr_lookup(lookup_path: Path, modified_ns: int) -> dict[str, int]:
    """Load player-to-UTR-profile-ID mappings from columns A and D."""
    lookup_df = backend.pd.read_excel(lookup_path, usecols=[0, 3])
    lookup_df.columns = ["Player", "UTR Profile ID"]
    lookup_df = lookup_df.dropna(subset=["Player", "UTR Profile ID"]).copy()
    lookup_df["Player"] = lookup_df["Player"].astype("string").str.strip()
    lookup_df["UTR Profile ID"] = backend.pd.to_numeric(
        lookup_df["UTR Profile ID"], errors="coerce"
    )
    return {
        player: int(profile_id)
        for player, profile_id in lookup_df.dropna(subset=["UTR Profile ID"]).itertuples(index=False)
        if player
    }


@st.cache_data(show_spinner=False)
def load_wtn_lookup(lookup_path: Path, modified_ns: int) -> dict[str, str]:
    """Load player-to-WTN-person-ID mappings from columns A and C."""
    lookup_df = backend.pd.read_excel(lookup_path, usecols=[0, 2])
    lookup_df.columns = ["Player", "WTN Person ID"]
    lookup_df = lookup_df.dropna(subset=["Player", "WTN Person ID"]).copy()
    lookup_df["Player"] = lookup_df["Player"].astype("string").str.strip()
    lookup_df["WTN Person ID"] = lookup_df["WTN Person ID"].astype("string").str.strip()
    lookup_df = lookup_df[
        lookup_df["WTN Person ID"].apply(
            lambda value: bool(WTN_OBJECT_ID_RE.match(str(value)))
        )
    ]
    return {
        player: str(wtn_id).lower()
        for player, wtn_id in lookup_df.itertuples(index=False)
        if player and wtn_id
    }


def fetch_h2h_statistics(left_id: str, right_id: str) -> tuple[dict | None, dict | None]:
    """Fetch H2H data and retain response diagnostics when the call fails."""
    request = Request(
        f"{H2H_BACKEND_URL}/h2h",
        data=json.dumps({"left_id": left_id, "right_id": right_id}).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-API-Key": H2H_BACKEND_API_KEY,
            "User-Agent": "H2H-Streamlit-Client/1.0",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            response_body = response.read().decode("utf-8", errors="replace")
            diagnostic = {
                "status": response.status,
                "reason": response.reason,
                "headers": dict(response.headers.items()),
                "body": response_body,
            }
    except HTTPError as exc:
        return None, {
            "status": exc.code,
            "reason": str(exc.reason),
            "headers": dict(exc.headers.items()) if exc.headers else {},
            "body": exc.read().decode("utf-8", errors="replace"),
        }
    except (URLError, TimeoutError, OSError) as exc:
        return None, {
            "error_type": type(exc).__name__,
            "message": str(exc),
        }

    try:
        return json.loads(response_body), diagnostic
    except json.JSONDecodeError as exc:
        diagnostic["json_error"] = str(exc)
        return None, diagnostic


def fetch_common_matches(left_id: str, right_id: str) -> tuple[dict | None, dict | None]:
    """Fetch match history for two WTN IDs through the authorized backend."""
    request = Request(
        H2H_MATCH_HISTORY_URL,
        data=json.dumps({"left_id": left_id, "right_id": right_id}).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-API-Key": H2H_BACKEND_API_KEY,
            "User-Agent": "H2H-Streamlit-Client/1.0",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            response_body = response.read().decode("utf-8", errors="replace")
            diagnostic = {
                "status": response.status,
                "reason": response.reason,
                "headers": dict(response.headers.items()),
                "body": response_body,
            }
    except HTTPError as exc:
        return None, {
            "status": exc.code,
            "reason": str(exc.reason),
            "headers": dict(exc.headers.items()) if exc.headers else {},
            "body": exc.read().decode("utf-8", errors="replace"),
        }
    except (URLError, TimeoutError, OSError) as exc:
        return None, {"error_type": type(exc).__name__, "message": str(exc)}

    try:
        return json.loads(response_body), diagnostic
    except json.JSONDecodeError as exc:
        diagnostic["json_error"] = str(exc)
        return None, diagnostic


def _match_history_rows(history_response: dict, left_name: str, right_name: str) -> list[dict[str, str]]:
    """Convert the common-match response into display-ready history rows."""
    if not isinstance(history_response, dict):
        return []
    data = history_response.get("data")
    statistics = data.get("matchUpStatistics") if isinstance(data, dict) else None
    match_ups = statistics.get("matchUps") if isinstance(statistics, dict) else None
    items = match_ups.get("items") if isinstance(match_ups, dict) else []
    if not isinstance(items, list):
        return []
    rows = []
    for match in items:
        if not isinstance(match, dict):
            continue
        sides = match.get("sides") or []
        side_names = []
        for side in sides:
            people = side.get("players") or []
            person = (people[0].get("person") or {}) if people else {}
            side_names.append(" ".join(
                part for part in [person.get("nativeGivenName"), person.get("nativeFamilyName")] if part
            ).strip())

        winning_side = match.get("winningSide")
        winner_index = {"SIDE1": 0, "SIDE2": 1}.get(winning_side)
        winner_name = side_names[winner_index] if winner_index is not None and winner_index < len(side_names) else "Unavailable"
        if winner_name not in {left_name, right_name}:
            winner_name = left_name if winner_name == "" and side_names[:1] else winner_name

        scores = []
        score = match.get("score") or {}
        for set_score in score.get("sets") or []:
            winner_games = set_score.get("winnerGamesWon")
            loser_games = set_score.get("loserGamesWon")
            if winner_games is None or loser_games is None:
                continue
            set_text = f"{winner_games}-{loser_games}"
            tiebreaker = set_score.get("tiebreaker") or {}
            if tiebreaker.get("winnerPointsWon") is not None and tiebreaker.get("loserPointsWon") is not None:
                set_text += f" [{tiebreaker['winnerPointsWon']}-{tiebreaker['loserPointsWon']}]"
            scores.append(set_text)
        super_tiebreak = score.get("superTiebreak") or {}
        if super_tiebreak.get("winnerPointsWon") is not None and super_tiebreak.get("loserPointsWon") is not None:
            scores.append(f"[{super_tiebreak['winnerPointsWon']}-{super_tiebreak['loserPointsWon']}]")

        start = str(match.get("start") or "")[:10]
        try:
            parsed_date = datetime.strptime(start, "%Y-%m-%d")
            match_date = f"{parsed_date.month}/{parsed_date.day}/{parsed_date.year}"
        except ValueError:
            match_date = start or "Unavailable"
        rows.append({
            "Date": match_date,
            "Tournament": str((match.get("tournament") or {}).get("name") or "Unavailable"),
            "Winner": winner_name,
            "Score": ",  ".join(scores) or "Unavailable",
        })
    return rows


def fetch_common_competitors(left_id: str, right_id: str) -> tuple[dict | None, dict | None]:
    """Fetch shared opponents for two WTN IDs through the authorized backend."""
    request = Request(
        H2H_COMMON_COMPETITORS_URL,
        data=json.dumps({"left_id": left_id, "right_id": right_id}).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "X-API-Key": H2H_BACKEND_API_KEY,
            "User-Agent": "H2H-Streamlit-Client/1.0",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            response_body = response.read().decode("utf-8", errors="replace")
            diagnostic = {
                "status": response.status,
                "reason": response.reason,
                "headers": dict(response.headers.items()),
                "body": response_body,
            }
    except HTTPError as exc:
        return None, {
            "status": exc.code,
            "reason": str(exc.reason),
            "headers": dict(exc.headers.items()) if exc.headers else {},
            "body": exc.read().decode("utf-8", errors="replace"),
        }
    except (URLError, TimeoutError, OSError) as exc:
        return None, {"error_type": type(exc).__name__, "message": str(exc)}

    try:
        return json.loads(response_body), diagnostic
    except json.JSONDecodeError as exc:
        diagnostic["json_error"] = str(exc)
        return None, diagnostic


def _invert_record(record: dict | None) -> dict:
    """Flip a shared-opponent record so wins and losses read from the compared player's side."""
    record = record or {}
    wins = record.get("losses")
    losses = record.get("wins")
    return {**record, "wins": wins, "losses": losses}


def _common_competitor_rows(response: dict) -> list[dict[str, object]]:
    """Convert common-opponent API data into display-ready rows."""
    competitors = (
        ((response.get("data") or {}).get("matchUpStatistics") or {})
        .get("commonCompetitors", [])
    )
    rows = []
    for competitor in competitors:
        single_wtn = next(
            (
                item.get("tennisNumber")
                for item in competitor.get("worldTennisNumbers") or []
                if str(item.get("type") or "").upper() == "SINGLE"
            ),
            None,
        )
        first_name = str(competitor.get("nativeGivenName") or "").strip()
        last_name = str(competitor.get("nativeFamilyName") or "").strip()
        name = " ".join(part for part in [first_name, last_name] if part)
        if first_name.upper() == "UNKNOWN" and last_name.upper() == "UNKNOWN":
            name = "Name Unavailable"
        rows.append({
            "id": str(competitor.get("id") or ""),
            "name": name or "Unavailable",
            "wtn": f"{float(single_wtn):.1f}" if isinstance(single_wtn, (int, float)) else "unavailable",
            "p1": _invert_record(competitor.get("p1Vs")),
            "p2": _invert_record(competitor.get("p2Vs")),
        })
    return rows


def _record_text(record: dict) -> str:
    wins = record.get("wins")
    losses = record.get("losses")
    return f"{wins}-{losses}" if isinstance(wins, int) and isinstance(losses, int) else "unavailable"


def _win_percentage(record: dict) -> str:
    wins = record.get("wins")
    total_matches = record.get("totalMatches")
    if not isinstance(wins, int) or not isinstance(total_matches, int) or total_matches <= 0:
        return "unavailable"
    return f"{wins / total_matches:.0%}"


def _common_competitor_totals(rows: list[dict[str, object]]) -> tuple[dict, dict]:
    """Sum wins and losses across the shared-opponent rows."""
    totals = []
    for side in ("p1", "p2"):
        wins = sum(
            record.get("wins", 0)
            for row in rows
            for record in [row[side]]
            if isinstance(record.get("wins"), int)
        )
        losses = sum(
            record.get("losses", 0)
            for row in rows
            for record in [row[side]]
            if isinstance(record.get("losses"), int)
        )
        totals.append({"wins": wins, "losses": losses, "totalMatches": wins + losses})
    return totals[0], totals[1]


def _full_name_from_person(person: dict) -> str:
    first = str(person.get("nativeGivenName") or "").strip()
    last = str(person.get("nativeFamilyName") or "").strip()
    name = " ".join(part for part in [first, last] if part)
    return name if name else "Unknown"


def _singles_wtn(person: dict) -> str:
    for item in person.get("worldTennisNumbers") or []:
        if str(item.get("type") or "").upper() == "SINGLE":
            value = item.get("tennisNumber")
            if isinstance(value, (int, float)):
                return f"{value:.1f}"
    return "unavailable"


def render_h2h_view(players: list[str], selected_player: str, rankings: pd.DataFrame) -> None:
    if not UTR_LOOKUP_WORKBOOK.exists():
        st.error(f"Could not find {UTR_LOOKUP_WORKBOOK.name}. Add it beside app3.py to use H2H.")
        return

    try:
        lookup = load_wtn_lookup(UTR_LOOKUP_WORKBOOK, UTR_LOOKUP_WORKBOOK.stat().st_mtime_ns)
    except (OSError, ValueError, backend.pd.errors.EmptyDataError) as exc:
        st.error(f"Could not load WTN ID mapping from list.xlsx: {exc}")
        return

    if not lookup:
        st.warning("No WTN IDs were found in list.xlsx column C.")
        return

    default_left = selected_player if selected_player in players else players[0]
    if "h2h_left_player" not in st.session_state or st.session_state["h2h_left_player"] not in players:
        st.session_state["h2h_left_player"] = default_left
    if "h2h_right_player" not in st.session_state or st.session_state["h2h_right_player"] not in players:
        st.session_state["h2h_right_player"] = next(
            (player for player in players if player != st.session_state["h2h_left_player"]),
            st.session_state["h2h_left_player"],
        )

    side_left, side_mid, side_right = st.columns([1, 3, 1])
    with side_mid:
        st.subheader("H2H (Head-to-Head)", anchor=False)
        st.caption("Compare two players using World Tennis Number matchup statistics.")
        with st.form("h2h_compare_form", clear_on_submit=False):
            left_col, right_col = st.columns(2)
            with left_col:
                left_player = st.selectbox(
                    "First Player",
                    players,
                    index=players.index(st.session_state["h2h_left_player"]),
                    key="h2h_left",
                )
            with right_col:
                right_player = st.selectbox(
                    "Second Player",
                    players,
                    index=players.index(st.session_state["h2h_right_player"]),
                    key="h2h_right",
                )
            compare_clicked = st.form_submit_button("Compare", type="primary")

    if compare_clicked:
        st.session_state.pop("h2h_failed_response", None)
        st.session_state.pop("h2h_history_failed_response", None)
        st.session_state.pop("h2h_competitors_failed_response", None)
        if left_player == right_player:
            st.warning("Select two different players for H2H comparison.")
        else:
            left_id = lookup.get(left_player)
            right_id = lookup.get(right_player)
            missing = [name for name, pid in [(left_player, left_id), (right_player, right_id)] if not pid]
            if missing:
                warning_left, warning_mid, warning_right = st.columns([1, 3, 1])
                with warning_mid:
                    st.warning("WTN data unavailable for one or both players")
            else:
                spinner_left, spinner_mid, spinner_right = st.columns([1, 3, 1])
                with spinner_mid:
                    with st.spinner("Fetching H2H statistics from World Tennis Number..."):
                        response, diagnostic = fetch_h2h_statistics(left_id, right_id)
                if not response:
                    st.session_state["h2h_failed_response"] = diagnostic
                    error_left, error_mid, error_right = st.columns([1, 3, 1])
                    with error_mid:
                        st.error("Unable to process the comparison.")
                elif response.get("errors"):
                    st.session_state["h2h_failed_response"] = {
                        **diagnostic,
                        "graphql_errors": response["errors"],
                        "graphql_response": response,
                    }
                    error_left, error_mid, error_right = st.columns([1, 3, 1])
                    with error_mid:
                        st.error("Unable to process the comparison.")
                else:
                    with st.spinner("Fetching match history..."):
                        match_history, history_diagnostic = fetch_common_matches(left_id, right_id)
                    with st.spinner("Fetching competitors in common..."):
                        common_competitors, competitors_diagnostic = fetch_common_competitors(left_id, right_id)
                    st.session_state["h2h_left_player"] = left_player
                    st.session_state["h2h_right_player"] = right_player
                    st.session_state["h2h_result"] = {
                        "left_player": left_player,
                        "right_player": right_player,
                        "left_id": left_id,
                        "right_id": right_id,
                        "response": response,
                        "match_history": match_history,
                        "common_competitors": common_competitors,
                    }
                    if match_history and match_history.get("errors"):
                        st.session_state["h2h_history_failed_response"] = {
                            **history_diagnostic,
                            "graphql_errors": match_history["errors"],
                            "graphql_response": match_history,
                        }
                    elif not match_history:
                        st.session_state["h2h_history_failed_response"] = history_diagnostic
                    if common_competitors and common_competitors.get("errors"):
                        st.session_state["h2h_competitors_failed_response"] = {
                            **competitors_diagnostic,
                            "graphql_errors": common_competitors["errors"],
                            "graphql_response": common_competitors,
                        }
                    elif not common_competitors:
                        st.session_state["h2h_competitors_failed_response"] = competitors_diagnostic

    failed_response = st.session_state.get("h2h_failed_response")
    if failed_response:
        diagnostic_left, diagnostic_mid, diagnostic_right = st.columns([1, 3, 1])
        with diagnostic_mid:
            with st.expander("Diagnostic Information", expanded=False):
                st.json(failed_response)

    if "h2h_result" not in st.session_state:
        info_left, info_mid, info_right = st.columns([1, 3, 1])
        with info_mid:
            st.info("Select two players and click Compare to load head-to-head statistics.")
        return

    h2h_result = st.session_state["h2h_result"]
    left_id = h2h_result["left_id"]
    right_id = h2h_result["right_id"]
    response = h2h_result["response"]
    data = response.get("data") or {}
    persons = ((data.get("persons") or {}).get("items") or [])
    persons_by_id = {str(person.get("id")): person for person in persons if person.get("id")}
    left_person = persons_by_id.get(left_id, {})
    right_person = persons_by_id.get(right_id, {})

    left_name = _full_name_from_person(left_person) if left_person else h2h_result["left_player"]
    right_name = _full_name_from_person(right_person) if right_person else h2h_result["right_player"]
    left_wtn = _singles_wtn(left_person)
    right_wtn = _singles_wtn(right_person)

    def ota_profile_url(player_name: str) -> str | None:
        player_rows = rankings.loc[rankings["Player"].eq(player_name)]
        return get_latest_profile_url(player_rows)

    def linked_name(player_name: str, profile_url: str | None = None) -> str:
        safe_name = escape(player_name)
        if not profile_url:
            return safe_name
        return (
            f'<a href="{escape(profile_url, quote=True)}" target="_blank" '
            f'rel="noopener noreferrer">{safe_name}</a>'
        )

    left_profile_url = ota_profile_url(h2h_result["left_player"])
    right_profile_url = ota_profile_url(h2h_result["right_player"])
    left_avatar_html = build_avatar_html(
        h2h_result["left_player"],
        rankings.loc[rankings["Player"].eq(h2h_result["left_player"])],
        "h2h-avatar",
    )
    right_avatar_html = build_avatar_html(
        h2h_result["right_player"],
        rankings.loc[rankings["Player"].eq(h2h_result["right_player"])],
        "h2h-avatar",
    )
    left_utr_rating = calculate_utr_rating(
        get_utr_profile_for_player(h2h_result["left_player"])
    )
    right_utr_rating = calculate_utr_rating(
        get_utr_profile_for_player(h2h_result["right_player"])
    )
    left_utr_text = "unavailable" if left_utr_rating is None else f"{left_utr_rating:.2f}"
    right_utr_text = "unavailable" if right_utr_rating is None else f"{right_utr_rating:.2f}"

    stats = (data.get("matchUpStatistics") or {}).get("matchUpVsStatisticsCount") or {}
    total_matches = stats.get("totalMatches")
    left_wins = stats.get("wins")
    right_wins = stats.get("losses")
    left_record = "unavailable"
    right_record = "unavailable"
    if isinstance(total_matches, int) and isinstance(left_wins, int) and isinstance(right_wins, int):
        left_record = f"{left_wins} of {total_matches}"
        right_record = f"{right_wins} of {total_matches}"

    left_likelihood_raw = (data.get("matchUpStatistics") or {}).get("winLikelihood")
    if isinstance(left_likelihood_raw, (int, float)):
        left_likelihood = round(float(left_likelihood_raw), 1)
        right_likelihood = round(100.0 - left_likelihood, 1)
        left_likelihood_text = f"{left_likelihood:.0f}%" if float(left_likelihood).is_integer() else f"{left_likelihood:.1f}%"
        right_likelihood_text = f"{right_likelihood:.0f}%" if float(right_likelihood).is_integer() else f"{right_likelihood:.1f}%"
    else:
        left_likelihood_text = "unavailable"
        right_likelihood_text = "unavailable"

    left_wins_display = left_wins if isinstance(left_wins, int) else "—"
    right_wins_display = right_wins if isinstance(right_wins, int) else "—"
    not_competed = isinstance(left_wins, int) and isinstance(right_wins, int) and left_wins == 0 and right_wins == 0

    mid_left, mid_col, mid_right = st.columns([1, 3, 1])
    with mid_col:
        st.html(
            f"""
            <div class="h2h-scoreband">
                <div class="h2h-scoreband__name h2h-scoreband__player">{left_avatar_html}<span>{linked_name(left_name, left_profile_url)}</span></div>
                <div>
                    <div class="h2h-scoreband__score">
                        <span class="h2h-scoreband__score-left">{left_wins_display}</span>
                        <span class="h2h-scoreband__vs">vs</span>
                        <span class="h2h-scoreband__score-right">{right_wins_display}</span>
                    </div>
                    <div class="h2h-scoreband__labels"><span>Win</span><span>Wins</span></div>
                    {"<div class='h2h-not-competed'>These players have not competed</div>" if not_competed else ""}
                </div>
                <div class="h2h-scoreband__name h2h-scoreband__name--right h2h-scoreband__player h2h-scoreband__player--right">{right_avatar_html}<span>{linked_name(right_name, right_profile_url)}</span></div>
            </div>
            """,
        )

        metrics_table_html = f"""
        <table class="h2h-metrics-table">
            <colgroup>
                <col style="width: 33.333%">
                <col style="width: 33.333%">
                <col style="width: 33.333%">
            </colgroup>
            <thead>
                <tr>
                    <th class="player-header--left">{linked_name(left_name, left_profile_url)}</th>
                    <th class="metric-header"></th>
                    <th class="player-header--right">{linked_name(right_name, right_profile_url)}</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td>{escape(left_wtn)}</td>
                    <td class="metric-label">Singles WTN</td>
                    <td>{escape(right_wtn)}</td>
                </tr>
                <tr>
                    <td>{escape(left_utr_text)}</td>
                    <td class="metric-label">Singles UTR</td>
                    <td>{escape(right_utr_text)}</td>
                </tr>
                <tr>
                    <td>{escape(left_likelihood_text)}</td>
                    <td class="metric-label">Singles Win Likelihood</td>
                    <td>{escape(right_likelihood_text)}</td>
                </tr>
            </tbody>
        </table>
        """
        st.markdown(metrics_table_html, unsafe_allow_html=True)

        history_response = h2h_result.get("match_history") or {}
        history_rows = _match_history_rows(history_response, left_name, right_name)
        st.html('<div class="h2h-section-title">Match History</div>')
        if history_rows:
            history_table_html = """
            <table class="h2h-history">
                <colgroup>
                    <col style="width: 14%">
                    <col style="width: 42%">
                    <col style="width: 22%">
                    <col style="width: 22%">
                </colgroup>
                <thead>
                    <tr><th>Date</th><th>Tournament</th><th>Winner</th><th>Score</th></tr>
                </thead>
                <tbody>
            """
            for row in history_rows:
                history_table_html += (
                    f"<tr><td>{escape(row['Date'])}</td>"
                    f"<td>{escape(row['Tournament'])}</td>"
                    f"<td class=\"h2h-history__winner\">{escape(row['Winner'])}</td>"
                    f"<td>{escape(row['Score'])}</td></tr>"
                )
            history_table_html += "</tbody></table>"
            st.markdown(history_table_html, unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="h2h-history__empty">No match history available.</div>',
                unsafe_allow_html=True,
            )

        history_failed_response = st.session_state.get("h2h_history_failed_response")
        if history_failed_response:
            with st.expander("Match history request diagnostics", expanded=False):
                st.json(history_failed_response)

        competitor_rows = _common_competitor_rows(h2h_result.get("common_competitors") or {})
        st.html('<div class="h2h-section-title">Competitors in Common</div>')
        if competitor_rows:
            competitor_totals_left, competitor_totals_right = _common_competitor_totals(competitor_rows)
            competitor_table_html = """
            <table class="h2h-common-competitors">
                <colgroup>
                    <col style="width: 34%">
                    <col style="width: 16.5%">
                    <col style="width: 16.5%">
                    <col style="width: 16.5%">
                    <col style="width: 16.5%">
                </colgroup>
                <thead>
                    <tr><th rowspan="2"></th>
                        <th class="h2h-common-competitors__left" colspan="2">{left_header}</th>
                        <th class="h2h-common-competitors__right" colspan="2">{right_header}</th></tr>
                    <tr><th>W / L</th><th>WIN %</th><th>W / L</th><th>WIN %</th></tr>
                </thead>
                <tbody>
            """.format(
                left_header=linked_name(left_name, left_profile_url),
                right_header=linked_name(right_name, right_profile_url),
            )
            for competitor in competitor_rows:
                competitor_url = ota_profile_url(competitor["name"])
                competitor_identity = (
                    f'{linked_name(competitor["name"], competitor_url)} '
                    f'<sup>[{escape(str(competitor["wtn"]))}]</sup>'
                )
                competitor_table_html += (
                    f'<tr><td class="h2h-common-competitors__name">{competitor_identity}</td>'
                    f'<td>{escape(_record_text(competitor["p1"]))}</td>'
                    f'<td>{escape(_win_percentage(competitor["p1"]))}</td>'
                    f'<td>{escape(_record_text(competitor["p2"]))}</td>'
                    f'<td>{escape(_win_percentage(competitor["p2"]))}</td></tr>'
                )
            competitor_table_html += (
                "</tbody><tfoot><tr>"
                "<th>TOTAL</th>"
                f"<td>{escape(_record_text(competitor_totals_left))}</td>"
                f"<td>{escape(_win_percentage(competitor_totals_left))}</td>"
                f"<td>{escape(_record_text(competitor_totals_right))}</td>"
                f"<td>{escape(_win_percentage(competitor_totals_right))}</td>"
                "</tr></tfoot></table>"
            )
            st.markdown(competitor_table_html, unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="h2h-history__empty">No competitors in common.</div>',
                unsafe_allow_html=True,
            )

        competitors_failed_response = st.session_state.get("h2h_competitors_failed_response")
        if competitors_failed_response:
            with st.expander("Competitors in common request diagnostics", expanded=False):
                st.json(competitors_failed_response)


@st.cache_data(show_spinner=False, ttl=21600)
def fetch_utr_profile(profile_id: int) -> dict | None:
    """Fetch a UTR Sports player profile using the configured request headers."""
    request = Request(
        UTR_PROFILE_URL.format(player_id=profile_id),
        headers={
            "Accept": "application/json",
            "Origin": "https://app.utrsports.net",
            "Referer": "https://app.utrsports.net/",
            "X-Client-Name": "buildId - 302800",
            "User-Agent": "Mozilla/5.0",
        },
        method="GET",
    )
    try:
        with urlopen(request, timeout=15) as response:
            return json.load(response)
    except (HTTPError, URLError, TimeoutError, OSError):
        return None


def calculate_utr_rating(profile: dict | None) -> float | None:
    """Calculate the current UTR from the three-month rating-change block."""
    if not profile:
        return None
    details = profile.get("threeMonthRatingChangeDetails") or {}
    rating = details.get("rating")
    difference = details.get("ratingDifference")
    if isinstance(rating, (int, float)) and isinstance(difference, (int, float)) and (rating or difference):
        return round(rating + difference, 2)
    # Some profiles report a zeroed change block and only carry the rating in singlesUtr.
    singles_utr = profile.get("singlesUtr")
    if isinstance(singles_utr, (int, float)) and singles_utr:
        return round(float(singles_utr), 2)
    return None


def get_utr_profile_for_player(selected_player: str) -> dict | None:
    """Resolve a player's UTR ID from the lookup workbook and fetch their profile."""
    if not UTR_LOOKUP_WORKBOOK.exists():
        return None
    try:
        profile_ids = load_utr_lookup(
            UTR_LOOKUP_WORKBOOK,
            UTR_LOOKUP_WORKBOOK.stat().st_mtime_ns,
        )
        profile_id = profile_ids.get(selected_player)
        return fetch_utr_profile(profile_id) if profile_id is not None else None
    except (OSError, ValueError, backend.pd.errors.EmptyDataError):
        return None


def render_sidebar_status(rankings, modified_ns: int) -> None:
    """Show low-emphasis dataset freshness and size at the sidebar bottom."""
    dataset_line, updated_line = backend.latest_dataset_text(
        rankings, backend.MASTER_WORKBOOK, modified_ns
    )
    st.sidebar.markdown(
        f"""
        <div class="sidebar-data-status">
            <div class="sidebar-data-status__title">Data status</div>
            <div>{escape(dataset_line)}</div>
            <div>{escape(updated_line)}</div>
            <div>{rankings['Week Label'].nunique():,} weeks | {rankings['Player'].nunique():,} players | {len(rankings):,} records</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_player_header(selected_player: str, player_df, utr_profile: dict | None) -> None:
    """Display selected-player identity with avatar and profile metadata."""
    latest = player_df.tail(1)
    if latest.empty:
        st.warning("No profile records are available for the selected player.")
        return

    row = latest.iloc[0]
    profile_url = backend.get_latest_profile_url(player_df)
    player_name = escape(selected_player)
    if profile_url:
        player_name = (
            f'<a href="{escape(profile_url, quote=True)}" target="_blank" '
            f'rel="noopener noreferrer">{player_name}</a>'
        )

    avatar_url = resolve_player_avatar_url(selected_player, player_df)
    avatar_initials = player_initials(selected_player)
    avatar_color_set = [
        ("#dfe7ee", "#b7c5cf"),
        ("#dce9f5", "#aec7dc"),
        ("#e7e3d7", "#c5c2b7"),
        ("#e4dce9", "#c7b5d1"),
    ]
    color_index = sum(ord(ch) for ch in selected_player) % len(avatar_color_set)
    start_color, end_color = avatar_color_set[color_index]

    if avatar_url:
        avatar_html = (
            f'<div class="player-header__avatar">'
            f'<img src="{escape(avatar_url, quote=True)}" alt="{escape(selected_player, quote=True)} avatar" />'
            f'</div>'
        )
    else:
        avatar_html = (
            f'<div class="player-header__avatar player-header__avatar--initials" '
            f'style="background: linear-gradient(135deg, {start_color} 0%, {end_color} 100%);">'
            f'{escape(avatar_initials)}</div>'
        )

    def detail(column: str) -> str:
        value = row.get(column)
        if backend.pd.isna(value) or not str(value).strip():
            return "unavailable"
        if column == "Year of Birth":
            return backend.format_value(column, value)
        return str(value).strip()

    def profile_value(field: str) -> str | None:
        value = (utr_profile or {}).get(field)
        if value is None or not str(value).strip():
            return None
        return str(value).strip()

    def title_case(value: str) -> str:
        return value.replace("_", " ").replace("-", " ").strip().title()

    descriptors = [
        f"Born {detail('Year of Birth')}",
        detail("Province"),
        detail("Club"),
    ]
    birthplace = profile_value("birthplace")
    if birthplace:
        descriptors.append(f"Birthplace: {birthplace}")
    height_inches = (utr_profile or {}).get("heightInches")
    if isinstance(height_inches, (int, float)) and height_inches > 0:
        total_inches = int(round(height_inches))
        descriptors.append(f"Height: {total_inches // 12}' {total_inches % 12}\"")
    dominant_hand = profile_value("dominantHand")
    if dominant_hand:
        descriptors.append(f"Hand: {title_case(dominant_hand)}")
    backhand = profile_value("backhand")
    if backhand:
        descriptors.append(f"{title_case(backhand)} backhand")
    racket_brand, racket_type = profile_value("racketBrand"), profile_value("racketType")
    if racket_brand or racket_type:
        descriptors.append(f"Racquet: {' '.join(value for value in [racket_brand, racket_type] if value)}")
    for field, label in [
        ("preferredBall", "Ball"),
        ("apparelBrand", "Apparel"),
        ("shoesBrand", "Shoes"),
    ]:
        value = profile_value(field)
        if value:
            descriptors.append(f"{label}: {value}")

    safe_descriptors = " &nbsp; | &nbsp; ".join(escape(descriptor) for descriptor in descriptors)

    st.markdown(
        f"""
        <div class="player-header">
            <div class="player-header__row">
                {avatar_html}
                <div>
                    <div class="player-header__name">{player_name}</div>
                    <div class="player-header__meta">{safe_descriptors}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_player_metrics(rankings, player_df, utr_profile: dict | None) -> None:
    """Restore the selected player's latest metrics above the active view."""
    latest = player_df.tail(1)
    if latest.empty:
        return

    row = latest.iloc[0]
    provincial_rank = backend.get_latest_provincial_rank(rankings, player_df.iloc[-1]["Player"])
    utr_rating = calculate_utr_rating(utr_profile)
    win_loss = row.get("Singles W/L-YTD")
    win_loss_display = "unavailable" if backend.pd.isna(win_loss) else str(win_loss)
    wins, losses = backend.parse_win_loss_ytd(win_loss)

    if wins is not None and losses is not None and wins + losses > 0:
        win_width = wins / (wins + losses) * 100
        loss_width = 100 - win_width
        win_loss_bar = (
            '<div class="player-metric__win-loss-bar">'
            f'<span class="player-metric__wins" style="width: {win_width:.4f}%"></span>'
            f'<span class="player-metric__losses" style="width: {loss_width:.4f}%"></span>'
            "</div>"
        )
    else:
        win_loss_bar = ""

    standard_metrics = [
        ("National rank", backend.format_value("Rank", row.get("Rank"))),
        ("Provincial rank", backend.format_value("Rank", provincial_rank)),
        ("Points", backend.format_value("Points", row.get("Points"))),
    ]

    standard_html = "".join(
        f'<div class="player-metric"><div class="player-metric__label">{escape(label)}</div>'
        f'<div class="player-metric__value-area"><div class="player-metric__value">{escape(value)}</div></div></div>'
        for label, value in standard_metrics
    )
    win_loss_html = (
        '<div class="player-metric player-metric--win-loss">'
        '<div class="player-metric__label">Singles W/L YTD</div>'
        f'<div class="player-metric__value-area"><div class="player-metric__value">{escape(win_loss_display)}</div>{win_loss_bar}</div>'
        "</div>"
    )
    wtn_value = row.get("Singles WTN")
    if backend.pd.isna(wtn_value):
        wtn_donut = ""
    else:
        filled_degrees = max(0.0, min(270.0, (40.0 - float(wtn_value)) / 39.0 * 270.0))
        wtn_donut = (
            f'<div class="player-metric__wtn-donut" style="--wtn-filled: {filled_degrees:.2f}deg"></div>'
        )
    wtn_html = (
        '<div class="player-metric player-metric--wtn">'
        '<div class="player-metric__label">WTN</div>'
        f'<div class="player-metric__value-area">{wtn_donut}'
        f'<div class="player-metric__value">{escape(backend.format_value("Singles WTN", wtn_value))}</div></div></div>'
    )
    if utr_rating is None:
        utr_donut = ""
        utr_display = "unavailable"
    else:
        filled_degrees = max(0.0, min(270.0, (utr_rating - 1.0) / 15.5 * 270.0))
        utr_donut = (
            f'<div class="player-metric__rating-donut" style="--wtn-filled: {filled_degrees:.2f}deg"></div>'
        )
        utr_display = f"{utr_rating:.2f}"
    utr_html = (
        '<div class="player-metric player-metric--utr">'
        '<div class="player-metric__label">UTR</div>'
        f'<div class="player-metric__value-area">{utr_donut}'
        f'<div class="player-metric__value">{utr_display}</div></div></div>'
    )
    metric_parts = [
        f'<div class="player-metric"><div class="player-metric__label">{escape("National rank")}</div>'
        f'<div class="player-metric__value-area"><div class="player-metric__value">{escape(backend.format_value("Rank", row.get("Rank")))}</div></div></div>',
        win_loss_html,
        f'<div class="player-metric"><div class="player-metric__label">{escape("Points")}</div>'
        f'<div class="player-metric__value-area"><div class="player-metric__value">{escape(backend.format_value("Points", row.get("Points")))}</div></div></div>',
        f'<div class="player-metric"><div class="player-metric__label">{escape("Provincial rank")}</div>'
        f'<div class="player-metric__value-area"><div class="player-metric__value">{escape(backend.format_value("Rank", provincial_rank))}</div></div></div>',
        wtn_html,
        utr_html,
    ]
    metric_html = "".join(metric_parts)
    st.markdown(f'<div class="player-metrics">{metric_html}</div>', unsafe_allow_html=True)


def build_player_pdf_report(selected_player: str, player_df: pd.DataFrame, analysis: str) -> bytes:
    """Create a coach-shareable PDF report with profile, analysis, history, and charts."""
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]
    body_style.leading = 14
    story = [Paragraph(f"{escape(selected_player)} | Player Performance Report", title_style)]

    latest = player_df.tail(1)
    if not latest.empty:
        row = latest.iloc[0]
        profile_rows = [
            ["Latest national rank", format_value("Rank", row.get("Rank"))],
            ["Ranking points", format_value("Points", row.get("Points"))],
            ["Singles WTN", format_value("Singles WTN", row.get("Singles WTN"))],
            ["Year of birth", format_value("Year of Birth", row.get("Year of Birth"))],
            ["Province", str(row.get("Province") or "unavailable")],
            ["Club", str(row.get("Club") or "unavailable")],
        ]
        profile_table = Table(profile_rows, colWidths=[1.7 * inch, 2.7 * inch])
        profile_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF1F8")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1F2937")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.extend([Spacer(1, 0.15 * inch), profile_table])

    story.extend([Spacer(1, 0.22 * inch), Paragraph("AI Performance Analysis", heading_style)])
    for paragraph in analysis.splitlines():
        text = paragraph.strip()
        if not text:
            continue
        if text.startswith("### "):
            story.extend([Spacer(1, 0.08 * inch), Paragraph(escape(text[4:]), heading_style)])
        else:
            formatted_text = convert_markdown_to_reportlab(text)
            story.append(Paragraph(formatted_text, body_style))

    story.extend([Spacer(1, 0.22 * inch), Paragraph("Player History", heading_style)])
    history_df = build_player_history_display(player_df).head(12)
    if not history_df.empty:
        history_columns = [column for column in ["Week", "Rank", "Points", "Singles WTN", "Singles W/L-YTD"] if column in history_df.columns]
        history_rows = [history_columns] + history_df[history_columns].astype(str).values.tolist()
        history_table = Table(history_rows, repeatRows=1)
        history_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(history_table)

    story.extend([Spacer(1, 0.22 * inch), Paragraph("Player History Chart", heading_style)])
    history_chart = build_pdf_line_chart(player_df, "Rank")
    if history_chart:
        story.append(history_chart)

    story.extend([Spacer(1, 0.16 * inch), Paragraph("Player Charts", heading_style)])
    for metric in ["Points", "Singles WTN", "Singles W/L-YTD %"]:
        if metric in player_df.columns:
            chart = build_pdf_line_chart(player_df, metric)
            if chart:
                story.extend([Spacer(1, 0.08 * inch), chart])

    document.build(story)
    return output.getvalue()


def render_ai_analysis(selected_player: str, player_df, rankings, modified_ns: int) -> None:
    """Generate and retain an AI response for the active player and dataset."""
    st.subheader(f"AI Analysis for {selected_player}", anchor=False)

    if "GEMINI_API_KEY" not in st.secrets:
        st.warning("Google Gemini API key not found. Add GEMINI_API_KEY to Streamlit secrets to use this view.")
        return

    response_key = f"ai_analysis:{selected_player}:{modified_ns}"
    if st.button("Generate AI Analysis", type="primary"):
        try:
            player_history = backend.build_player_history_display(player_df).to_string()
        except Exception:
            player_history = player_df.to_string()

        try:
            ytd_df, _, _ = backend.calculate_ytd_trends(rankings, player_df)
            ytd_summary = ytd_df.to_string()
        except ValueError:
            ytd_summary = "YTD trends data unavailable because the player is missing from a comparison week."

        prompt = f"""
        You are an expert tennis analyst. Analyze the following statistics for the youth tennis player '{selected_player}'.

        1. Player Match & Ranking History:
        {player_history}

        2. Year-to-Date Trends:
        {ytd_summary}

        Provide a concise, professional performance summary under these headings:
        ### Overall Performance Overview
        ### Key Strengths & Progress
        ### Areas for Improvement
        ### Outlook & Recommendations
        """
        with st.spinner("Analyzing player performance with Google Gemini..."):
            try:
                client = backend.genai.Client(api_key=st.secrets["GEMINI_API_KEY"])
                response = client.models.generate_content(
                    model="gemini-3.6-flash", contents=prompt
                )
                st.session_state[response_key] = response.text
            except Exception as exc:
                st.error(f"Failed to generate AI analysis: {exc}")

    saved_response = st.session_state.get(response_key)
    if saved_response:
        st.markdown(saved_response)
        try:
            report = build_player_pdf_report(selected_player, player_df, saved_response)
            st.download_button(
                "Download coach report (PDF)",
                data=report,
                file_name=f"{selected_player.lower().replace(' ', '_')}_performance_report.pdf",
                mime="application/pdf",
            )
        except Exception as exc:
            st.warning(f"The analysis is available, but the PDF report could not be generated: {exc}")
    else:
        st.info("Generate an AI Analysis of this player, with a report that can be saved as a PDF")


def render_view(view: str, rankings, players, selected_player: str | None, player_df, modified_ns: int) -> None:
    """Render the active navigation view with explicit data-state messaging."""
    if selected_player is None and view in PLAYER_VIEWS:
        st.info("Select a player in the sidebar to load this view.")
        return
    if view == "Player history":
        if player_df.empty:
            st.warning("No historical records are available for the selected player.")
            return
        st.caption("Most recent week appears first. Numeric cells include week-to-week trend arrows.")
        display_df = backend.build_player_history_display(player_df)
        if display_df.empty:
            st.info("The selected player has a record, but no displayable history metrics are available.")
            return
        st.dataframe(
            backend.style_trend_table(display_df, backend.NUMERIC_TREND_COLUMNS),
            width="stretch",
            height=backend.table_auto_height(len(display_df)),
            hide_index=True,
        )
    elif view == "Player Charts":
        st.subheader("Player Comparison Charts", anchor=False)
        chart_players = st.multiselect("Select players to compare", players, default=[selected_player])
        if not chart_players:
            st.info("Select at least one player to display chart data.")
            return
        chart_rankings = rankings.loc[rankings["Player"].isin(chart_players)]
        if chart_rankings.empty:
            st.warning("No ranking records are available for the selected chart players.")
            return
        available_metrics = [metric for metric in backend.OPTIONAL_CHART_METRICS if metric in rankings.columns]
        if not available_metrics:
            st.warning("This workbook does not contain numeric metrics that can be charted.")
            return
        for metric in available_metrics:
            backend.chart_multi_player_metric(chart_rankings, chart_players, metric)
    elif view == "YTD Player trends":
        try:
            trend_df, oldest_label, latest_label = backend.calculate_ytd_trends(rankings, player_df)
        except ValueError as exc:
            st.warning(str(exc))
            st.info("YTD comparison requires the selected player to appear in both the oldest and latest weekly datasets.")
            return
        if trend_df.empty:
            st.info("The player exists in both comparison weeks, but no comparable numeric metrics are available.")
            return
        st.caption(f"Comparison: {oldest_label} to {latest_label}")
        card_metrics = ["Rank", "Points", "Tournaments", "Singles WTN", "Career Matches", "Singles W/L-YTD %"]
        card_df = trend_df.loc[trend_df["Metric"].isin(card_metrics)]
        for start in range(0, len(card_df), 3):
            columns = st.columns(3)
            for column, (_, row) in zip(columns, card_df.iloc[start:start + 3].iterrows()):
                column.metric(row["Metric"], row[f"Latest ({latest_label})"], delta=row["Metric card delta"])
        st.dataframe(trend_df.drop(columns=["Metric card delta"]), width="stretch", hide_index=True)
    elif view == "Top movers":
        try:
            moved_up, moved_down, previous_label, latest_label = backend.build_top_movers(rankings)
        except ValueError as exc:
            st.warning(str(exc))
            return
        st.caption(f"Ranking movement: {previous_label} to {latest_label}")
        if moved_up.empty and moved_down.empty:
            st.info("No players changed ranking position between the two latest weeks.")
            return
        if not moved_up.empty:
            st.subheader("Top 10 ranking movers up", anchor=False)
            render_player_dataframe(backend.style_trend_table(moved_up, ["Rank Trend"] + backend.TOP_MOVER_METRIC_COLUMNS), rankings, width="stretch", hide_index=True)
        else:
            st.info("No players improved their rank between the two latest weeks.")
        if not moved_down.empty:
            st.subheader("Top 10 ranking movers down", anchor=False)
            render_player_dataframe(backend.style_trend_table(moved_down, ["Rank Trend"] + backend.TOP_MOVER_METRIC_COLUMNS), rankings, width="stretch", hide_index=True)
        else:
            st.info("No players declined in rank between the two latest weeks.")
    elif view == "Latest Week Trends":
        try:
            trend_df, previous_label, latest_label = backend.build_latest_week_trends(rankings)
        except ValueError as exc:
            st.warning(str(exc))
            return
        if trend_df.empty:
            st.info("No comparable player records are available for the latest two weeks.")
            return
        st.caption(f"Comparison: {previous_label} to {latest_label}. Only players listed in {latest_label} are included.")
        render_player_dataframe(backend.style_trend_table(trend_df, backend.LATEST_TRENDS_COLUMNS), rankings, width="stretch", height=520, hide_index=True)
    elif view == "Top 100":
        backend.render_top_n_tab(rankings, 100)
    elif view == "Top 50":
        backend.render_top_n_tab(rankings, 50)
    elif view == "Top 20":
        backend.render_top_n_tab(rankings, 20)
    elif view == "Top 50 Ontario":
        backend.render_top_50_ontario_tab(rankings)
    elif view == "AI Analysis":
        render_ai_analysis(selected_player, player_df, rankings, modified_ns)
    elif view == "Multi-Player Download":
        backend.render_multi_player_download_tab(rankings)
    elif view == "Head-to-Head":
        render_h2h_view(players, selected_player, rankings)


def main() -> None:
    apply_styles()
    st.title("OTA Girls U14 Ranking Dashboard")

    if not backend.MASTER_WORKBOOK.exists():
        st.error(f"Could not find {backend.MASTER_WORKBOOK.name}. Place it beside app.py and rerun the app.")
        return

    modified_ns = backend.MASTER_WORKBOOK.stat().st_mtime_ns
    try:
        rankings = backend.load_rankings(backend.MASTER_WORKBOOK, modified_ns)
    except Exception as exc:
        st.error(f"Could not load workbook: {exc}")
        return

    players = sorted(rankings["Player"].dropna().unique())
    if not players:
        st.error("No players were found in the weekly worksheets.")
        return

    with st.sidebar:
        st.header("Dashboard")
        player_choice = st.selectbox("Player", [PLAYER_PLACEHOLDER, *players], index=0)
        section = st.selectbox("Section", list(VIEWS_BY_SECTION))
        view = st.radio("View", VIEWS_BY_SECTION[section])
        st.divider()
        st.caption("Use the player selector to update player-specific views.")

    render_sidebar_status(rankings, modified_ns)
    selected_player = None if player_choice == PLAYER_PLACEHOLDER else player_choice

    if selected_player is None:
        player_df = rankings.iloc[0:0]
    else:
        player_df = rankings.loc[rankings["Player"].eq(selected_player)].sort_values("Week Sort")
        utr_profile = get_utr_profile_for_player(selected_player)
        render_player_header(selected_player, player_df, utr_profile)
        render_player_metrics(rankings, player_df, utr_profile)

    render_view(view, rankings, players, selected_player, player_df, modified_ns)


if __name__ == "__main__":
    main()
